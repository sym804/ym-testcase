import io
from urllib.parse import quote

from fastapi.responses import StreamingResponse
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter


def export_testcases_excel(project, testcases, split_sheets: bool) -> StreamingResponse:
    """TC 목록을 Excel 파일로 생성하여 StreamingResponse로 반환한다."""
    wb = Workbook()

    # -- Styles --
    dark_fill = PatternFill(start_color="2F3136", end_color="2F3136", fill_type="solid")
    header_font = Font(name="Malgun Gothic", bold=True, color="FFFFFF", size=10)
    cell_font = Font(name="Malgun Gothic", size=10)
    thin_border = Border(
        left=Side(style="thin"),
        right=Side(style="thin"),
        top=Side(style="thin"),
        bottom=Side(style="thin"),
    )
    center_align = Alignment(horizontal="center", vertical="center", wrap_text=True)
    left_align = Alignment(horizontal="left", vertical="center", wrap_text=True)

    pass_fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
    fail_fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
    block_fill = PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid")
    na_fill = PatternFill(start_color="D9D9D9", end_color="D9D9D9", fill_type="solid")

    headers = [
        "No", "TC ID", "Type", "Category", "Depth1", "Depth2",
        "Priority", "Platform", "Precondition", "Steps",
        "Expected Result", "Remarks",
        "Result", "Actual Result", "Issue Link", "Remarks (Run)",
    ]
    col_widths = [6, 10, 10, 15, 18, 18, 10, 12, 25, 35, 35, 20, 10, 35, 20, 20]

    def _write_sheet(ws, sheet_title, tcs, show_title=True):
        """시트 하나에 TC 목록을 쓴다."""
        if show_title:
            max_col_letter = get_column_letter(len(headers) + 1)
            ws.merge_cells(f"B1:{max_col_letter}1")
            title_cell = ws["B1"]
            title_cell.value = f"{project.name} - {sheet_title}"
            title_cell.font = Font(name="Malgun Gothic", bold=True, size=14)
            title_cell.alignment = Alignment(horizontal="center", vertical="center")

            ws.merge_cells(f"B3:{max_col_letter}3")
            ws["B3"].value = f"Project: {project.name}"
            ws["B3"].font = Font(name="Malgun Gothic", size=10)

        for idx, (header, width) in enumerate(zip(headers, col_widths)):
            col = idx + 2
            cell = ws.cell(row=5, column=col, value=header)
            cell.font = header_font
            cell.fill = dark_fill
            cell.alignment = center_align
            cell.border = thin_border
            ws.column_dimensions[get_column_letter(col)].width = width

        for row_offset, tc in enumerate(tcs):
            row = 6 + row_offset
            values = [
                tc.no, tc.tc_id, tc.type, tc.category, tc.depth1, tc.depth2,
                tc.priority, tc.test_type, tc.precondition, tc.test_steps,
                tc.expected_result, tc.remarks,
                "", "", "", "",
            ]
            for col_offset, value in enumerate(values):
                col = col_offset + 2
                cell = ws.cell(row=row, column=col, value=value)
                cell.font = cell_font
                cell.border = thin_border
                if col_offset < 2 or col_offset == 6 or col_offset == 7 or col_offset == 12 or col_offset == 14:
                    cell.alignment = center_align
                else:
                    cell.alignment = left_align

        result_col = 14 + 2  # Result column
        for row in range(6, 6 + len(tcs)):
            cell = ws.cell(row=row, column=result_col)
            val = str(cell.value or "").upper()
            if val == "PASS":
                cell.fill = pass_fill
            elif val == "FAIL":
                cell.fill = fail_fill
            elif val == "BLOCK":
                cell.fill = block_fill
            elif val in ("NA", "N/A"):
                cell.fill = na_fill

        ws.freeze_panes = "B6"

    if split_sheets:
        # 시트별 분리: sheet_name 기준으로 엑셀 탭 생성
        from collections import OrderedDict
        sheets_map: OrderedDict[str, list] = OrderedDict()
        for tc in testcases:
            name = tc.sheet_name or "기본"
            sheets_map.setdefault(name, []).append(tc)

        wb.remove(wb.active)  # 기본 빈 시트 제거
        for sheet_name, tcs in sheets_map.items():
            # 엑셀 시트명은 31자 제한, 특수문자 제거
            safe_name = sheet_name[:31].replace("/", "-").replace("\\", "-").replace("*", "").replace("?", "").replace("[", "").replace("]", "").replace(":", "")
            ws = wb.create_sheet(title=safe_name)
            _write_sheet(ws, sheet_name, tcs)
    else:
        # 통합: 단일 시트
        ws = wb.active
        ws.title = "Test Cases"
        _write_sheet(ws, "Test Cases", testcases)

    output = io.BytesIO()
    wb.save(output)
    output.seek(0)

    filename = f"{project.name}_TestCases.xlsx"
    encoded = quote(filename)
    return StreamingResponse(
        output,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename*=UTF-8''{encoded}"},
    )

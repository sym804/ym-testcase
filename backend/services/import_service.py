import csv
import io
import logging
import re
from typing import Optional

from fastapi import HTTPException, UploadFile
from sqlalchemy import func
from sqlalchemy.orm import Session
from openpyxl import load_workbook

from models import TestCase

logger = logging.getLogger(__name__)


# Expected header mapping (column letter -> field name). The template uses
# headers in row 5, columns B-S.
HEADER_MAP = {
    # ── English headers ──
    "No": "no",
    "no": "no",
    "TC ID": "tc_id",
    "TC_ID": "tc_id",
    "tc_id": "tc_id",
    "Type": "type",
    "type": "type",
    "Category": "category",
    "category": "category",
    "Depth1": "depth1",
    "depth1": "depth1",
    "Depth 1": "depth1",
    "Depth2": "depth2",
    "depth2": "depth2",
    "Depth 2": "depth2",
    "Depth 3": "depth3",
    "Depth3": "depth3",
    "Priority": "priority",
    "priority": "priority",
    "Test Type": "test_type",
    "test_type": "test_type",
    "Platform": "test_type",
    "Precondition": "precondition",
    "precondition": "precondition",
    "Precondition /\nTest Data": "precondition",
    "Precondition / Test Data": "precondition",
    "Steps": "test_steps",
    "steps": "test_steps",
    "Test Steps": "test_steps",
    "Expected Result": "expected_result",
    "expected_result": "expected_result",
    "Actual Result": "remarks",
    "R1": "r1",
    "R2": "r2",
    "R3": "r3",
    "Issue Link": "_skip",
    "issue_link": "_skip",
    "Assignee": "_skip",
    "assignee": "_skip",
    "Remarks": "remarks",
    "remarks": "remarks",
    "Requirement\nID": "remarks",
    "Requirement ID": "remarks",
    # ── Korean headers ──
    "대분류": "category",
    "분류": "category",
    "카테고리": "category",
    "구분": "type",
    "유형": "type",
    "우선순위": "priority",
    "우선 순위": "priority",
    "플랫폼": "test_type",
    "자동화 구현 여부": "test_type",
    "사전조건": "precondition",
    "사전조건/\n테스트 데이터": "precondition",
    "사전조건/ 테스트 데이터": "precondition",
    "사전조건/테스트 데이터": "precondition",
    "테스트 케이스": "test_steps",
    "테스트 절차": "test_steps",
    "테스트 스텝": "test_steps",
    "체크 포인트": "test_steps",
    "체크포인트": "test_steps",
    "기대 결과": "expected_result",
    "기대결과": "expected_result",
    "예상 결과": "expected_result",
    "테스트 결과": "_skip",  # result columns (round-specific), skip
    "이슈": "_skip",
    "버그/이슈": "_skip",
    "확인자": "_skip",
    "담당자": "_skip",
    "비고": "remarks",
    "사전조건/테스트 가이드": "precondition",
    "심각도": "remarks",
    "결과": "_skip",
    "자동화": "_skip",
}


def _resolve_merged(ws, row: int, col: int):
    """Return effective value for a cell that may be part of a merged range."""
    cell = ws.cell(row=row, column=col)
    if cell.value is not None:
        return cell.value
    for merged_range in ws.merged_cells.ranges:
        if cell.coordinate in merged_range:
            return ws.cell(row=merged_range.min_row, column=merged_range.min_col).value
    return None


SKIP_SHEETS = {"DASHBOARD", "요약", "SUMMARY"}


def _detect_header_row(ws) -> Optional[int]:
    """시트에서 헤더 행을 자동 감지한다."""
    for row_idx in range(1, min(ws.max_row + 1, 20)):
        for col_idx in range(1, ws.max_column + 1):
            val = ws.cell(row=row_idx, column=col_idx).value
            if val and str(val).strip() == "TC ID":
                return row_idx
    # Fallback: look for any known header
    for row_idx in range(1, min(ws.max_row + 1, 20)):
        for col_idx in range(1, ws.max_column + 1):
            val = ws.cell(row=row_idx, column=col_idx).value
            if val and str(val).strip() in HEADER_MAP:
                return row_idx
    return None


def _count_tc_rows(ws, header_row: int) -> int:
    """헤더 아래에서 유효한 TC 행 수를 센다 (섹션 헤더 제외)."""
    count = 0
    # 전체 컬럼 매핑 구축
    all_cols: dict[int, str] = {}
    tc_id_col: Optional[int] = None
    no_col: Optional[int] = None
    for col_idx in range(1, ws.max_column + 1):
        val = ws.cell(row=header_row, column=col_idx).value
        if val and str(val).strip() in HEADER_MAP:
            mapped = HEADER_MAP[str(val).strip()]
            if mapped == "_skip":
                continue
            all_cols[col_idx] = mapped
            if mapped == "tc_id":
                tc_id_col = col_idx
            elif mapped == "no":
                no_col = col_idx

    empty_streak = 0
    for row_idx in range(header_row + 1, ws.max_row + 1):
        # 머지된 섹션 헤더 감지: 여러 컬럼 값이 모두 동일
        vals = set()
        for col_idx in all_cols:
            v = _resolve_merged(ws, row_idx, col_idx)
            if v is not None and str(v).strip():
                vals.add(str(v).strip())
        if len(vals) == 0:
            empty_streak += 1
            if empty_streak >= 10:
                break
            continue
        if len(vals) <= 1 and len(all_cols) > 1:
            continue  # 섹션 헤더

        # No가 숫자인지 확인
        has_valid = False
        if tc_id_col:
            tc_val = _resolve_merged(ws, row_idx, tc_id_col)
            if tc_val and str(tc_val).strip():
                has_valid = True
        if not has_valid and no_col:
            no_val = _resolve_merged(ws, row_idx, no_col)
            if no_val is not None:
                try:
                    int(no_val)
                    has_valid = True
                except (ValueError, TypeError):
                    pass

        if has_valid:
            count += 1
            empty_streak = 0
        else:
            empty_streak += 1
            if empty_streak >= 10:
                break
    return count


def _parse_sheet(ws, project_id: int, user_id: int, db: Session, no_offset: int = 0, sheet_name: str = "기본") -> dict:
    """단일 시트를 파싱하여 TC를 DB에 추가/업데이트한다. {"created": N, "updated": N} 반환."""
    header_row = _detect_header_row(ws)
    if header_row is None:
        return 0

    col_map: dict[int, str] = {}
    for col_idx in range(1, ws.max_column + 1):
        val = ws.cell(row=header_row, column=col_idx).value
        if val and str(val).strip() in HEADER_MAP:
            mapped = HEADER_MAP[str(val).strip()]
            if mapped == "_skip":
                continue
            col_map[col_idx] = mapped

    created_count = 0
    updated_count = 0
    empty_streak = 0

    # 기존 TC 맵 (같은 프로젝트+시트+tc_id → TC 객체)
    existing_tcs = (
        db.query(TestCase)
        .filter(TestCase.project_id == project_id, TestCase.sheet_name == sheet_name, TestCase.deleted_at.is_(None))
        .all()
    )
    existing_map = {tc.tc_id: tc for tc in existing_tcs}

    for row_idx in range(header_row + 1, ws.max_row + 1):
        row_data: dict = {}
        for col_idx, field in col_map.items():
            value = _resolve_merged(ws, row_idx, col_idx)
            if value is not None:
                parsed = str(value).strip() if not isinstance(value, (int, float)) else value
                if field in row_data and row_data[field]:
                    continue
                row_data[field] = parsed

        if not row_data or all(v == "" or v is None for v in row_data.values()):
            empty_streak += 1
            if empty_streak >= 10:
                break
            continue
        empty_streak = 0

        # 섹션 헤더 행 감지
        unique_vals = set(str(v).strip() for v in row_data.values() if v and str(v).strip())
        if len(unique_vals) <= 1 and len(row_data) > 1:
            continue

        # depth3 → depth2 병합
        if row_data.get("depth3"):
            if not row_data.get("depth2"):
                row_data["depth2"] = row_data["depth3"]
            else:
                row_data["depth2"] = f"{row_data['depth2']} > {row_data['depth3']}"
        row_data.pop("depth3", None)

        if "tc_id" not in row_data or not row_data.get("tc_id"):
            if "no" in row_data:
                try:
                    no_val = int(row_data["no"])
                    row_data["tc_id"] = f"TC-{no_val:03d}"
                except (ValueError, TypeError):
                    continue
            else:
                continue

        if "no" not in row_data:
            row_data["no"] = no_offset + created_count + updated_count + 1
        else:
            try:
                row_data["no"] = int(row_data["no"]) + no_offset
            except (ValueError, TypeError):
                row_data["no"] = no_offset + created_count + updated_count + 1

        tc_id_val = str(row_data.get("tc_id", ""))
        fields = dict(
            no=row_data.get("no", no_offset + created_count + updated_count + 1),
            type=row_data.get("type"),
            category=row_data.get("category"),
            depth1=row_data.get("depth1"),
            depth2=row_data.get("depth2"),
            priority=row_data.get("priority"),
            test_type=row_data.get("test_type"),
            precondition=row_data.get("precondition"),
            test_steps=row_data.get("test_steps"),
            expected_result=row_data.get("expected_result"),
            r1=row_data.get("r1"),
            r2=row_data.get("r2"),
            r3=row_data.get("r3"),
            remarks=row_data.get("remarks"),
        )

        existing = existing_map.get(tc_id_val)
        if existing:
            # 덮어쓰기
            for key, val in fields.items():
                if val is not None:
                    setattr(existing, key, val)
            updated_count += 1
        else:
            tc = TestCase(
                project_id=project_id,
                created_by=user_id,
                tc_id=tc_id_val,
                sheet_name=sheet_name,
                **fields,
            )
            db.add(tc)
            created_count += 1

    return {"created": created_count, "updated": updated_count}


# ── Jira/Xray/Zephyr CSV 헤더 매핑 ───────────────────────────────────────────
JIRA_HEADER_MAP = {
    "Summary": "depth2",
    "Issue key": "tc_id",
    "Issue Key": "tc_id",
    "Key": "tc_id",
    "Issue id": "_skip",
    "Issue Type": "type",
    "Issue type": "type",
    "Priority": "priority",
    "Description": "test_steps",
    "Status": "_skip",
    "Component/s": "category",
    "Components": "category",
    "Component": "category",
    "Labels": "category",
    "Label": "category",
    "Assignee": "_skip",
    "Reporter": "_skip",
    "Fix Version/s": "remarks",
    "Fix Version": "remarks",
    "Affects Version/s": "_skip",
    "Created": "_skip",
    "Updated": "_skip",
    "Resolution": "_skip",
    "Epic Link": "depth1",
    "Epic Name": "depth1",
    "Sprint": "_skip",
    "Custom field (Test Steps)": "test_steps",
    "Test Step": "test_steps",
    "Test Data": "precondition",
    "Expected Result": "expected_result",
    "Precondition": "precondition",
    "Folder": "depth1",
    "Objective": "precondition",
    # Xray
    "Generic Test Definition": "test_steps",
    # Zephyr Scale
    "Name": "depth2",
    "Preconditions": "precondition",
    "Test Script (Step-by-Step) - Step": "test_steps",
    "Test Script (Step-by-Step) - Expected Result": "expected_result",
    "Test Script (Step-by-Step) - Test Data": "precondition",
}


def _parse_csv(file_content: bytes, project_id: int, user_id: int, db: Session, sheet_name: str = "CSV Import") -> dict:
    """CSV 파일을 파싱하여 TC를 DB에 추가/업데이트한다."""
    # 인코딩 감지
    text = None
    for enc in ["utf-8-sig", "utf-8", "cp949", "euc-kr", "latin-1"]:
        try:
            text = file_content.decode(enc)
            break
        except (UnicodeDecodeError, LookupError):
            continue
    if text is None:
        return {"created": 0, "updated": 0}

    reader = csv.DictReader(io.StringIO(text))
    if not reader.fieldnames:
        return {"created": 0, "updated": 0}

    # 컬럼 매핑 (HEADER_MAP + JIRA_HEADER_MAP 모두 사용)
    combined_map = {**HEADER_MAP, **JIRA_HEADER_MAP}
    col_map = {}
    for header in reader.fieldnames:
        h = header.strip()
        if h in combined_map:
            mapped = combined_map[h]
            if mapped != "_skip":
                col_map[header] = mapped

    if not col_map:
        return {"created": 0, "updated": 0}

    # 기존 TC 맵
    existing_tcs = (
        db.query(TestCase)
        .filter(TestCase.project_id == project_id, TestCase.sheet_name == sheet_name, TestCase.deleted_at.is_(None))
        .all()
    )
    existing_map = {tc.tc_id: tc for tc in existing_tcs}

    created_count = 0
    updated_count = 0

    for row_num, row in enumerate(reader, start=1):
        row_data = {}
        for csv_header, field in col_map.items():
            val = row.get(csv_header, "").strip()
            if val and field not in row_data:
                row_data[field] = val

        if not row_data or all(not v for v in row_data.values()):
            continue

        # tc_id 없으면 no 기반 생성
        if not row_data.get("tc_id"):
            if row_data.get("no"):
                try:
                    row_data["tc_id"] = f"TC-{int(row_data['no']):03d}"
                except (ValueError, TypeError):
                    row_data["tc_id"] = f"CSV-{row_num:04d}"
            else:
                row_data["tc_id"] = f"CSV-{row_num:04d}"

        if not row_data.get("no"):
            row_data["no"] = row_num

        try:
            row_data["no"] = int(row_data["no"])
        except (ValueError, TypeError):
            row_data["no"] = row_num

        # depth3 → depth2 병합
        if row_data.get("depth3"):
            if not row_data.get("depth2"):
                row_data["depth2"] = row_data["depth3"]
            else:
                row_data["depth2"] = f"{row_data['depth2']} > {row_data['depth3']}"
        row_data.pop("depth3", None)

        tc_id_val = str(row_data["tc_id"])
        fields = {k: row_data.get(k) for k in [
            "no", "type", "category", "depth1", "depth2", "priority",
            "test_type", "precondition", "test_steps", "expected_result",
            "r1", "r2", "r3", "remarks",
        ] if row_data.get(k) is not None}

        existing = existing_map.get(tc_id_val)
        if existing:
            for key, val in fields.items():
                if val is not None:
                    setattr(existing, key, val)
            updated_count += 1
        else:
            tc = TestCase(
                project_id=project_id,
                created_by=user_id,
                tc_id=tc_id_val,
                sheet_name=sheet_name,
                **fields,
            )
            db.add(tc)
            created_count += 1

    return {"created": created_count, "updated": updated_count}


MAX_IMPORT_SIZE = 10 * 1024 * 1024  # 10MB


def _load_workbook_from_upload(file: UploadFile):
    """UploadFile을 읽어 openpyxl Workbook으로 반환한다."""
    if not file.filename or not file.filename.endswith((".xlsx", ".xls")):
        raise HTTPException(status_code=400, detail="Only .xlsx files are supported")

    file_content = file.file.read()
    if len(file_content) > MAX_IMPORT_SIZE:
        raise HTTPException(
            status_code=413,
            detail=f"File too large. Maximum size is {MAX_IMPORT_SIZE // (1024*1024)}MB",
        )

    try:
        return load_workbook(filename=io.BytesIO(file_content), data_only=True)
    except Exception:
        logger.exception("Excel read failed")
        raise HTTPException(status_code=400, detail="Failed to read Excel file")


def _is_csv_file(filename: str) -> bool:
    return filename and filename.lower().endswith(".csv")


def _is_md_file(filename: str) -> bool:
    return filename and filename.lower().endswith(".md")


# ── Markdown Import ──────────────────────────────────────────────────────────

_MD_SEPARATOR_RE = re.compile(r"^\|[\s:-]+\|[\s:|-]*$")


def _decode_text(file_content: bytes) -> Optional[str]:
    """바이트를 텍스트로 디코딩한다 (CSV/MD 공용)."""
    for enc in ["utf-8-sig", "utf-8", "cp949", "euc-kr", "latin-1"]:
        try:
            return file_content.decode(enc)
        except (UnicodeDecodeError, LookupError):
            continue
    return None


def _parse_md_row(line: str) -> list[str]:
    """Markdown 테이블 행을 파싱하여 셀 리스트를 반환한다.
    이스케이프된 파이프(\\|)를 올바르게 처리한다."""
    line = line.strip()
    if line.startswith("|"):
        line = line[1:]
    if line.endswith("|"):
        line = line[:-1]
    # \| 를 플레이스홀더로 치환 후 split, 다시 복원
    placeholder = "\x00PIPE\x00"
    line = line.replace("\\|", placeholder)
    return [cell.strip().replace(placeholder, "|") for cell in line.split("|")]


def _parse_md_tables(file_content: bytes) -> list[dict]:
    """Markdown 파일에서 테이블들을 추출한다.
    반환: [{"name": str, "headers": list[str], "rows": list[list[str]]}, ...]
    """
    text = _decode_text(file_content)
    if not text:
        return []

    tables = []
    current_heading = None
    headers = None
    rows = []
    in_table = False
    unnamed_count = 0

    def _finalize_table():
        nonlocal headers, rows, in_table, unnamed_count
        if headers and rows:
            if current_heading:
                name = current_heading
            else:
                unnamed_count += 1
                name = f"MD Import" if unnamed_count == 1 else f"MD Import {unnamed_count}"
            # 중복 이름 처리
            existing_names = [t["name"] for t in tables]
            if name in existing_names:
                suffix = 2
                while f"{name} ({suffix})" in existing_names:
                    suffix += 1
                name = f"{name} ({suffix})"
            tables.append({"name": name, "headers": headers, "rows": rows})
        headers = None
        rows = []
        in_table = False

    for line in text.split("\n"):
        stripped = line.strip()

        # 헤딩 감지
        if stripped.startswith("#"):
            if in_table:
                _finalize_table()
            heading_text = stripped.lstrip("#").strip()
            if heading_text:
                current_heading = heading_text
            continue

        # 파이프 행 감지
        is_pipe_row = "|" in stripped and stripped.startswith("|")

        if is_pipe_row:
            # 구분자 행 (|---|---| 등) 스킵
            if _MD_SEPARATOR_RE.match(stripped):
                continue

            cells = _parse_md_row(stripped)

            if not in_table:
                # 첫 파이프 행 = 헤더
                headers = cells
                rows = []
                in_table = True
            else:
                rows.append(cells)
        else:
            # 테이블 밖의 줄
            if in_table:
                _finalize_table()

    # 파일 끝에 테이블이 남아있으면 저장
    if in_table:
        _finalize_table()

    return tables


def _preview_md(file_content: bytes, project_id: int, db: Session) -> list:
    """Markdown 파일의 미리보기 — 테이블별 TC 수와 기존 중복 수 반환."""
    tables = _parse_md_tables(file_content)
    result = []
    for t in tables:
        row_count = len(t["rows"])
        if row_count == 0:
            continue
        existing = db.query(TestCase).filter(
            TestCase.project_id == project_id,
            TestCase.sheet_name == t["name"],
            TestCase.deleted_at.is_(None),
        ).count()
        result.append({"name": t["name"], "tc_count": row_count, "existing": existing})
    return result


def _parse_md_table(table: dict, project_id: int, user_id: int, db: Session, sheet_name: str) -> dict:
    """파싱된 Markdown 테이블 하나를 TC로 DB에 추가/업데이트한다."""
    combined_map = {**HEADER_MAP, **JIRA_HEADER_MAP}

    # 헤더 → 필드 매핑 (인덱스 기반)
    col_map = {}
    for idx, h in enumerate(table["headers"]):
        h_stripped = h.strip()
        if h_stripped in combined_map:
            mapped = combined_map[h_stripped]
            if mapped != "_skip":
                col_map[idx] = mapped

    if not col_map:
        return {"created": 0, "updated": 0}

    # 기존 TC 맵
    existing_tcs = (
        db.query(TestCase)
        .filter(TestCase.project_id == project_id, TestCase.sheet_name == sheet_name, TestCase.deleted_at.is_(None))
        .all()
    )
    existing_map = {tc.tc_id: tc for tc in existing_tcs}

    created_count = 0
    updated_count = 0

    for row_num, cells in enumerate(table["rows"], start=1):
        row_data = {}
        for idx, field in col_map.items():
            if idx < len(cells):
                val = cells[idx].strip()
                if val and field not in row_data:
                    row_data[field] = val

        if not row_data or all(not v for v in row_data.values()):
            continue

        # tc_id 없으면 no 기반 생성
        if not row_data.get("tc_id"):
            if row_data.get("no"):
                try:
                    row_data["tc_id"] = f"TC-{int(row_data['no']):03d}"
                except (ValueError, TypeError):
                    row_data["tc_id"] = f"MD-{row_num:04d}"
            else:
                row_data["tc_id"] = f"MD-{row_num:04d}"

        if not row_data.get("no"):
            row_data["no"] = row_num

        try:
            row_data["no"] = int(row_data["no"])
        except (ValueError, TypeError):
            row_data["no"] = row_num

        # depth3 → depth2 병합
        if row_data.get("depth3"):
            if not row_data.get("depth2"):
                row_data["depth2"] = row_data["depth3"]
            else:
                row_data["depth2"] = f"{row_data['depth2']} > {row_data['depth3']}"
        row_data.pop("depth3", None)

        tc_id_val = str(row_data["tc_id"])
        fields = {k: row_data.get(k) for k in [
            "no", "type", "category", "depth1", "depth2", "priority",
            "test_type", "precondition", "test_steps", "expected_result",
            "r1", "r2", "r3", "remarks",
        ] if row_data.get(k) is not None}

        existing = existing_map.get(tc_id_val)
        if existing:
            for key, val in fields.items():
                if val is not None:
                    setattr(existing, key, val)
            updated_count += 1
        else:
            tc = TestCase(
                project_id=project_id,
                created_by=user_id,
                tc_id=tc_id_val,
                sheet_name=sheet_name,
                **fields,
            )
            db.add(tc)
            created_count += 1

    return {"created": created_count, "updated": updated_count}


def _preview_csv(file_content: bytes, project_id: int, db: Session) -> list:
    """CSV 파일의 미리보기 (시트 1개로 취급)."""
    text = None
    for enc in ["utf-8-sig", "utf-8", "cp949", "euc-kr", "latin-1"]:
        try:
            text = file_content.decode(enc)
            break
        except (UnicodeDecodeError, LookupError):
            continue
    if not text:
        return []

    reader = csv.DictReader(io.StringIO(text))
    row_count = sum(1 for _ in reader)
    if row_count == 0:
        return []

    sheet_name = "CSV Import"
    existing = db.query(TestCase).filter(
        TestCase.project_id == project_id,
        TestCase.sheet_name == sheet_name,
        TestCase.deleted_at.is_(None),
    ).count()
    return [{"name": sheet_name, "tc_count": row_count, "existing": existing}]

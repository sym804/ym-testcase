import io
import csv
import logging
from typing import List, Optional

from fastapi import APIRouter, Depends, HTTPException, status, UploadFile, File, Query
from pydantic import BaseModel
from fastapi.responses import StreamingResponse
from sqlalchemy import func
from sqlalchemy.orm import Session
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

from database import get_db
from models import User, Project, TestCase, TestCaseHistory
from schemas import (
    TestCaseCreate, TestCaseUpdate, TestCaseResponse, TestCaseBulkUpdate,
)
from auth import get_current_user, role_required, check_project_access

logger = logging.getLogger(__name__)

MAX_IMPORT_SIZE = 10 * 1024 * 1024  # 10MB


def _record_history(db: Session, tc: TestCase, changes: dict, user_id: int):
    """Record field-level change history for a test case."""
    for field, (old_val, new_val) in changes.items():
        db.add(TestCaseHistory(
            test_case_id=tc.id,
            changed_by=user_id,
            field_name=field,
            old_value=str(old_val) if old_val is not None else None,
            new_value=str(new_val) if new_val is not None else None,
        ))

router = APIRouter(
    prefix="/api/projects/{project_id}/testcases",
    tags=["testcases"],
)


def _get_project_or_404(project_id: int, db: Session) -> Project:
    project = db.query(Project).filter(Project.id == project_id).first()
    if not project:
        raise HTTPException(status_code=404, detail="Project not found")
    return project


# ── List ──────────────────────────────────────────────────────────────────────

@router.get("", response_model=List[TestCaseResponse])
def list_testcases(
    project_id: int,
    category: Optional[str] = Query(None),
    priority: Optional[str] = Query(None),
    search: Optional[str] = Query(None),
    sheet_name: Optional[str] = Query(None),
    db: Session = Depends(get_db),
    current_user: User = Depends(check_project_access("viewer")),
):
    _get_project_or_404(project_id, db)

    q = db.query(TestCase).filter(TestCase.project_id == project_id, TestCase.deleted_at.is_(None))

    if sheet_name:
        q = q.filter(TestCase.sheet_name == sheet_name)
    if category:
        q = q.filter(TestCase.category == category)
    if priority:
        q = q.filter(TestCase.priority == priority)
    if search:
        like = f"%{search}%"
        q = q.filter(
            (TestCase.tc_id.ilike(like))
            | (TestCase.depth1.ilike(like))
            | (TestCase.depth2.ilike(like))
            | (TestCase.test_steps.ilike(like))
            | (TestCase.expected_result.ilike(like))
        )

    return q.order_by(TestCase.no).all()


def _build_sheet_tree(sheets, tc_counts):
    """flat 시트 리스트를 트리 구조로 변환"""
    node_map = {}
    for s in sheets:
        node_map[s.id] = {
            "id": s.id,
            "name": s.name,
            "parent_id": s.parent_id,
            "sort_order": s.sort_order,
            "is_folder": getattr(s, "is_folder", False),
            "tc_count": tc_counts.get(s.name, 0),
            "children": [],
        }

    roots = []
    for s in sheets:
        node = node_map[s.id]
        if s.parent_id and s.parent_id in node_map:
            node_map[s.parent_id]["children"].append(node)
        else:
            roots.append(node)

    return roots


def _collect_descendant_names(sheet_id: int, db: Session, project_id: int) -> list[str]:
    """시트와 모든 하위 시트의 이름 목록 반환"""
    from models import TestCaseSheet
    names = []
    stack = [sheet_id]
    while stack:
        sid = stack.pop()
        s = db.query(TestCaseSheet).filter(TestCaseSheet.id == sid, TestCaseSheet.project_id == project_id).first()
        if s:
            names.append(s.name)
            children = db.query(TestCaseSheet).filter(
                TestCaseSheet.parent_id == sid, TestCaseSheet.project_id == project_id
            ).all()
            stack.extend(c.id for c in children)
    return names


@router.get("/sheets")
def list_sheets(
    project_id: int,
    flat: Optional[str] = Query(None),
    db: Session = Depends(get_db),
    current_user: User = Depends(check_project_access("viewer")),
):
    """프로젝트의 시트 목록을 트리 구조로 반환한다. ?flat=true 시 기존 flat 포맷."""
    _get_project_or_404(project_id, db)
    from sqlalchemy import func
    from models import TestCaseSheet

    # DB에 등록된 시트 목록
    registered = (
        db.query(TestCaseSheet)
        .filter(TestCaseSheet.project_id == project_id)
        .order_by(TestCaseSheet.sort_order, TestCaseSheet.id)
        .all()
    )

    # TC 기준 시트별 카운트
    tc_counts = dict(
        db.query(TestCase.sheet_name, func.count(TestCase.id))
        .filter(TestCase.project_id == project_id, TestCase.deleted_at.is_(None))
        .group_by(TestCase.sheet_name)
        .all()
    )

    if not registered:
        # 시트 테이블이 비어있으면 TC 기반으로 반환 (기존 호환)
        if not tc_counts:
            return []
        rows = (
            db.query(TestCase.sheet_name, func.count(TestCase.id), func.min(TestCase.id))
            .filter(TestCase.project_id == project_id, TestCase.deleted_at.is_(None))
            .group_by(TestCase.sheet_name)
            .order_by(func.min(TestCase.id))
            .all()
        )
        return [{"id": 0, "name": name or "기본", "parent_id": None, "sort_order": i, "tc_count": cnt, "children": []} for i, (name, cnt, _) in enumerate(rows)]

    # flat=true: 기존 호환 포맷
    if flat == "true":
        result = [{"name": s.name, "tc_count": tc_counts.get(s.name, 0)} for s in registered]
        registered_names = {s.name for s in registered}
        for name, cnt in tc_counts.items():
            if name and name not in registered_names:
                result.append({"name": name, "tc_count": cnt})
        return result

    # 트리 구조 반환
    tree = _build_sheet_tree(registered, tc_counts)

    # 등록 안 된 시트가 TC에 존재하면 root에 추가 (호환성)
    registered_names = {s.name for s in registered}
    for name, cnt in tc_counts.items():
        if name and name not in registered_names:
            tree.append({"id": 0, "name": name, "parent_id": None, "sort_order": 999, "tc_count": cnt, "children": []})

    return tree


class _SheetCreate(BaseModel):
    name: str
    parent_id: Optional[int] = None
    is_folder: bool = False


@router.post("/sheets")
def create_sheet(
    project_id: int,
    payload: _SheetCreate,
    db: Session = Depends(get_db),
    current_user: User = Depends(check_project_access("admin")),
):
    """빈 시트를 생성한다. parent_id로 부모 시트 지정 가능."""
    _get_project_or_404(project_id, db)
    from models import TestCaseSheet

    name = payload.name.strip()
    if not name:
        raise HTTPException(status_code=400, detail="시트 이름을 입력해 주세요.")

    exists = db.query(TestCaseSheet).filter(
        TestCaseSheet.project_id == project_id, TestCaseSheet.name == name
    ).first()
    if exists:
        raise HTTPException(status_code=400, detail="이미 존재하는 시트 이름입니다.")

    # parent_id 유효성 검증
    if payload.parent_id is not None:
        parent = db.query(TestCaseSheet).filter(
            TestCaseSheet.id == payload.parent_id, TestCaseSheet.project_id == project_id
        ).first()
        if not parent:
            raise HTTPException(status_code=400, detail="부모 시트를 찾을 수 없습니다.")
        # 부모가 일반 시트면 자동으로 폴더로 승격
        if not parent.is_folder:
            parent.is_folder = True

    # 같은 부모 아래에서 최대 sort_order
    sibling_q = db.query(TestCaseSheet.sort_order).filter(
        TestCaseSheet.project_id == project_id,
        TestCaseSheet.parent_id == payload.parent_id,
    )
    max_order = sibling_q.order_by(TestCaseSheet.sort_order.desc()).first()
    next_order = (max_order[0] + 1) if max_order else 0

    sheet = TestCaseSheet(project_id=project_id, name=name, sort_order=next_order, parent_id=payload.parent_id, is_folder=payload.is_folder)
    db.add(sheet)
    db.commit()
    db.refresh(sheet)
    return {"id": sheet.id, "name": sheet.name, "parent_id": sheet.parent_id, "sort_order": sheet.sort_order, "is_folder": sheet.is_folder, "tc_count": 0, "children": []}


class _SheetRename(BaseModel):
    new_name: str


@router.put("/sheets/{sheet_id}/rename")
def rename_sheet(
    project_id: int,
    sheet_id: int,
    payload: _SheetRename,
    db: Session = Depends(get_db),
    current_user: User = Depends(check_project_access("admin")),
):
    """시트 이름을 변경한다."""
    _get_project_or_404(project_id, db)
    from models import TestCaseSheet

    sheet = db.query(TestCaseSheet).filter(
        TestCaseSheet.id == sheet_id, TestCaseSheet.project_id == project_id
    ).first()
    if not sheet:
        raise HTTPException(status_code=404, detail="시트를 찾을 수 없습니다.")

    new_name = payload.new_name.strip()
    if not new_name:
        raise HTTPException(status_code=400, detail="시트 이름을 입력해 주세요.")

    dup = db.query(TestCaseSheet).filter(
        TestCaseSheet.project_id == project_id, TestCaseSheet.name == new_name, TestCaseSheet.id != sheet_id
    ).first()
    if dup:
        raise HTTPException(status_code=400, detail="이미 존재하는 시트 이름입니다.")

    old_name = sheet.name
    sheet.name = new_name

    # TC의 sheet_name도 업데이트
    db.query(TestCase).filter(
        TestCase.project_id == project_id, TestCase.sheet_name == old_name
    ).update({TestCase.sheet_name: new_name}, synchronize_session="fetch")

    db.commit()
    return {"id": sheet.id, "name": sheet.name, "old_name": old_name}


class _SheetMove(BaseModel):
    parent_id: Optional[int] = None
    sort_order: Optional[int] = None


@router.put("/sheets/{sheet_id}/move")
def move_sheet(
    project_id: int,
    sheet_id: int,
    payload: _SheetMove,
    db: Session = Depends(get_db),
    current_user: User = Depends(check_project_access("admin")),
):
    """시트를 다른 부모 아래로 이동하거나 순서를 변경한다."""
    _get_project_or_404(project_id, db)
    from models import TestCaseSheet

    sheet = db.query(TestCaseSheet).filter(
        TestCaseSheet.id == sheet_id, TestCaseSheet.project_id == project_id
    ).first()
    if not sheet:
        raise HTTPException(status_code=404, detail="시트를 찾을 수 없습니다.")

    # 자기 자신이나 자손을 부모로 설정하면 순환 참조
    if payload.parent_id is not None:
        if payload.parent_id == sheet_id:
            raise HTTPException(status_code=400, detail="자기 자신을 부모로 설정할 수 없습니다.")
        desc_names = _collect_descendant_names(sheet_id, db, project_id)
        parent = db.query(TestCaseSheet).filter(
            TestCaseSheet.id == payload.parent_id, TestCaseSheet.project_id == project_id
        ).first()
        if not parent:
            raise HTTPException(status_code=400, detail="부모 시트를 찾을 수 없습니다.")
        if parent.name in desc_names:
            raise HTTPException(status_code=400, detail="하위 시트를 부모로 설정할 수 없습니다.")

    sheet.parent_id = payload.parent_id
    if payload.sort_order is not None:
        sheet.sort_order = payload.sort_order
    db.commit()
    return {"id": sheet.id, "name": sheet.name, "parent_id": sheet.parent_id, "sort_order": sheet.sort_order}


@router.delete("/sheets/{sheet_name}")
def delete_sheet(
    project_id: int,
    sheet_name: str,
    db: Session = Depends(get_db),
    current_user: User = Depends(check_project_access("admin")),
):
    """시트와 속한 모든 TC를 삭제한다. 하위 시트도 함께 삭제."""
    _get_project_or_404(project_id, db)
    from models import TestCaseSheet, now_kst
    now = now_kst()

    # 시트 찾기
    sheet = db.query(TestCaseSheet).filter(
        TestCaseSheet.project_id == project_id, TestCaseSheet.name == sheet_name
    ).first()

    if sheet:
        # 하위 시트 포함 모든 이름 수집
        all_names = _collect_descendant_names(sheet.id, db, project_id)
    else:
        all_names = [sheet_name]

    # TC 소프트 삭제 (모든 하위 시트 포함)
    count = 0
    for name in all_names:
        c = (
            db.query(TestCase)
            .filter(TestCase.project_id == project_id, TestCase.sheet_name == name, TestCase.deleted_at.is_(None))
            .update({TestCase.deleted_at: now}, synchronize_session="fetch")
        )
        count += c

    # 시트 레코드 삭제 (하위 시트 포함, leaf부터 역순 삭제)
    if sheet:
        # 모든 하위 시트 ID 수집 후 역순 삭제 (leaf first)
        all_ids = []
        stack = [sheet.id]
        while stack:
            sid = stack.pop()
            all_ids.append(sid)
            children = db.query(TestCaseSheet).filter(
                TestCaseSheet.parent_id == sid, TestCaseSheet.project_id == project_id
            ).all()
            stack.extend(c.id for c in children)
        for sid in reversed(all_ids):
            db.query(TestCaseSheet).filter(TestCaseSheet.id == sid).delete()

    db.commit()
    return {"deleted": count, "sheet": sheet_name}


# ── 시트 무결성 검증 ─────────────────────────────────────────────────────────

def _validate_sheet_name(project_id: int, sheet_name: str | None, db: Session, *, auto_create_default: bool = False):
    """시트 이름 유효성 검증: 폴더 직접 추가 금지, 존재하지 않는 시트 금지.
    auto_create_default=True이면 "기본" 시트가 없을 때 자동 생성."""
    if not sheet_name:
        return
    from models import TestCaseSheet
    sheet = db.query(TestCaseSheet).filter(
        TestCaseSheet.project_id == project_id,
        TestCaseSheet.name == sheet_name,
    ).first()
    if not sheet:
        if auto_create_default and sheet_name == "기본":
            max_order = db.query(func.max(TestCaseSheet.sort_order)).filter(
                TestCaseSheet.project_id == project_id
            ).scalar() or 0
            new_sheet = TestCaseSheet(
                project_id=project_id, name="기본", sort_order=max_order + 1,
            )
            db.add(new_sheet)
            db.flush()
            return
        raise HTTPException(status_code=400, detail=f"존재하지 않는 시트입니다: {sheet_name}")
    if sheet.is_folder:
        raise HTTPException(status_code=400, detail="폴더에는 TC를 직접 추가할 수 없습니다. 하위 시트를 사용하세요.")


# ── Create ────────────────────────────────────────────────────────────────────

@router.post("", response_model=TestCaseResponse, status_code=status.HTTP_201_CREATED)
def create_testcase(
    project_id: int,
    payload: TestCaseCreate,
    db: Session = Depends(get_db),
    current_user: User = Depends(check_project_access("admin")),
):
    _get_project_or_404(project_id, db)
    _validate_sheet_name(project_id, payload.sheet_name, db, auto_create_default=True)

    tc = TestCase(
        project_id=project_id,
        created_by=current_user.id,
        **payload.model_dump(),
    )
    db.add(tc)
    db.commit()
    db.refresh(tc)
    return tc


# ── Bulk Update ───────────────────────────────────────────────────────────────
# NOTE: /bulk must be defined BEFORE /{tc_id} so FastAPI doesn't try to parse
# "bulk" as an integer tc_id.

@router.put("/bulk", response_model=List[TestCaseResponse])
def bulk_update_testcases(
    project_id: int,
    payload: TestCaseBulkUpdate,
    db: Session = Depends(get_db),
    current_user: User = Depends(check_project_access("admin")),
):
    _get_project_or_404(project_id, db)

    updated: list[TestCase] = []
    for item in payload.items:
        tc = db.query(TestCase).filter(
            TestCase.id == item.id, TestCase.project_id == project_id
        ).first()
        if not tc:
            continue
        item_data = item.model_dump(exclude_unset=True, exclude={"id"})
        if "sheet_name" in item_data:
            _validate_sheet_name(project_id, item_data["sheet_name"], db)
        changes = {}
        for key, value in item_data.items():
            old_val = getattr(tc, key)
            if str(old_val) != str(value):
                changes[key] = (old_val, value)
            setattr(tc, key, value)
        if changes:
            _record_history(db, tc, changes, current_user.id)
        updated.append(tc)

    db.commit()
    for tc in updated:
        db.refresh(tc)
    return updated


# ── Update ────────────────────────────────────────────────────────────────────

@router.put("/{tc_id}", response_model=TestCaseResponse)
def update_testcase(
    project_id: int,
    tc_id: int,
    payload: TestCaseUpdate,
    db: Session = Depends(get_db),
    current_user: User = Depends(check_project_access("admin")),
):
    _get_project_or_404(project_id, db)

    tc = db.query(TestCase).filter(
        TestCase.id == tc_id, TestCase.project_id == project_id
    ).first()
    if not tc:
        raise HTTPException(status_code=404, detail="Test case not found")

    update_data = payload.model_dump(exclude_unset=True)
    if "sheet_name" in update_data:
        _validate_sheet_name(project_id, update_data["sheet_name"], db)

    changes = {}
    for key, value in update_data.items():
        old_val = getattr(tc, key)
        if str(old_val) != str(value):
            changes[key] = (old_val, value)
        setattr(tc, key, value)

    if changes:
        _record_history(db, tc, changes, current_user.id)

    db.commit()
    db.refresh(tc)
    return tc


# ── Delete (soft) ────────────────────────────────────────────────────────────

@router.delete("/bulk")
def bulk_delete_testcases(
    project_id: int,
    ids: str = Query(..., description="쉼표 구분 TC ID 목록"),
    db: Session = Depends(get_db),
    current_user: User = Depends(check_project_access("admin")),
):
    """여러 TC를 한 번에 소프트 삭제한다."""
    _get_project_or_404(project_id, db)
    from models import now_kst
    id_list = [int(x.strip()) for x in ids.split(",") if x.strip()]
    now = now_kst()
    count = (
        db.query(TestCase)
        .filter(TestCase.id.in_(id_list), TestCase.project_id == project_id, TestCase.deleted_at.is_(None))
        .update({TestCase.deleted_at: now}, synchronize_session="fetch")
    )
    db.commit()
    return {"deleted": count}


@router.delete("/{tc_id}", status_code=status.HTTP_204_NO_CONTENT)
def delete_testcase(
    project_id: int,
    tc_id: int,
    db: Session = Depends(get_db),
    current_user: User = Depends(check_project_access("admin")),
):
    _get_project_or_404(project_id, db)

    tc = db.query(TestCase).filter(
        TestCase.id == tc_id, TestCase.project_id == project_id
    ).first()
    if not tc:
        raise HTTPException(status_code=404, detail="Test case not found")

    from models import now_kst
    tc.deleted_at = now_kst()
    db.commit()


# ── Restore ──────────────────────────────────────────────────────────────────

@router.post("/{tc_id}/restore", response_model=TestCaseResponse)
def restore_testcase(
    project_id: int,
    tc_id: int,
    db: Session = Depends(get_db),
    current_user: User = Depends(check_project_access("admin")),
):
    _get_project_or_404(project_id, db)

    tc = db.query(TestCase).filter(
        TestCase.id == tc_id, TestCase.project_id == project_id,
        TestCase.deleted_at.isnot(None),
    ).first()
    if not tc:
        raise HTTPException(status_code=404, detail="Test case not found or not deleted")

    tc.deleted_at = None
    db.commit()
    db.refresh(tc)
    return tc


# ── Import Excel ──────────────────────────────────────────────────────────────

# Expected header mapping (column letter -> field name). The template uses
# headers in row 5, columns B-S.
HEADER_MAP = {
    # ── English headers ──
    "No": "no",
    "no": "no",
    "TC ID": "tc_id",
    "TC_ID": "tc_id",
    "tc_id": "tc_id",
    "Type": "type",
    "type": "type",
    "Category": "category",
    "category": "category",
    "Depth1": "depth1",
    "depth1": "depth1",
    "Depth 1": "depth1",
    "Depth2": "depth2",
    "depth2": "depth2",
    "Depth 2": "depth2",
    "Depth 3": "depth3",
    "Depth3": "depth3",
    "Priority": "priority",
    "priority": "priority",
    "Test Type": "test_type",
    "test_type": "test_type",
    "Platform": "test_type",
    "Precondition": "precondition",
    "precondition": "precondition",
    "Precondition /\nTest Data": "precondition",
    "Precondition / Test Data": "precondition",
    "Steps": "test_steps",
    "steps": "test_steps",
    "Test Steps": "test_steps",
    "Expected Result": "expected_result",
    "expected_result": "expected_result",
    "Actual Result": "remarks",
    "R1": "r1",
    "R2": "r2",
    "R3": "r3",
    "Issue Link": "issue_link",
    "issue_link": "issue_link",
    "Assignee": "assignee",
    "assignee": "assignee",
    "Remarks": "remarks",
    "remarks": "remarks",
    "Requirement\nID": "remarks",
    "Requirement ID": "remarks",
    # ── Korean headers ──
    "대분류": "category",
    "분류": "category",
    "카테고리": "category",
    "구분": "type",
    "유형": "type",
    "우선순위": "priority",
    "우선 순위": "priority",
    "플랫폼": "test_type",
    "자동화 구현 여부": "test_type",
    "사전조건": "precondition",
    "사전조건/\n테스트 데이터": "precondition",
    "사전조건/ 테스트 데이터": "precondition",
    "사전조건/테스트 데이터": "precondition",
    "테스트 케이스": "test_steps",
    "테스트 절차": "test_steps",
    "테스트 스텝": "test_steps",
    "체크 포인트": "test_steps",
    "체크포인트": "test_steps",
    "기대 결과": "expected_result",
    "기대결과": "expected_result",
    "예상 결과": "expected_result",
    "테스트 결과": "_skip",  # result columns (round-specific), skip
    "이슈": "issue_link",
    "버그/이슈": "issue_link",
    "확인자": "assignee",
    "담당자": "assignee",
    "비고": "remarks",
    "사전조건/테스트 가이드": "precondition",
    "심각도": "remarks",
    "결과": "_skip",
    "자동화": "_skip",
}


def _resolve_merged(ws, row: int, col: int):
    """Return effective value for a cell that may be part of a merged range."""
    cell = ws.cell(row=row, column=col)
    if cell.value is not None:
        return cell.value
    for merged_range in ws.merged_cells.ranges:
        if cell.coordinate in merged_range:
            return ws.cell(row=merged_range.min_row, column=merged_range.min_col).value
    return None


SKIP_SHEETS = {"DASHBOARD", "요약", "SUMMARY"}


def _detect_header_row(ws) -> Optional[int]:
    """시트에서 헤더 행을 자동 감지한다."""
    for row_idx in range(1, min(ws.max_row + 1, 20)):
        for col_idx in range(1, ws.max_column + 1):
            val = ws.cell(row=row_idx, column=col_idx).value
            if val and str(val).strip() == "TC ID":
                return row_idx
    # Fallback: look for any known header
    for row_idx in range(1, min(ws.max_row + 1, 20)):
        for col_idx in range(1, ws.max_column + 1):
            val = ws.cell(row=row_idx, column=col_idx).value
            if val and str(val).strip() in HEADER_MAP:
                return row_idx
    return None


def _count_tc_rows(ws, header_row: int) -> int:
    """헤더 아래에서 유효한 TC 행 수를 센다 (섹션 헤더 제외)."""
    count = 0
    # 전체 컬럼 매핑 구축
    all_cols: dict[int, str] = {}
    tc_id_col: Optional[int] = None
    no_col: Optional[int] = None
    for col_idx in range(1, ws.max_column + 1):
        val = ws.cell(row=header_row, column=col_idx).value
        if val and str(val).strip() in HEADER_MAP:
            mapped = HEADER_MAP[str(val).strip()]
            if mapped == "_skip":
                continue
            all_cols[col_idx] = mapped
            if mapped == "tc_id":
                tc_id_col = col_idx
            elif mapped == "no":
                no_col = col_idx

    empty_streak = 0
    for row_idx in range(header_row + 1, ws.max_row + 1):
        # 머지된 섹션 헤더 감지: 여러 컬럼 값이 모두 동일
        vals = set()
        for col_idx in all_cols:
            v = _resolve_merged(ws, row_idx, col_idx)
            if v is not None and str(v).strip():
                vals.add(str(v).strip())
        if len(vals) == 0:
            empty_streak += 1
            if empty_streak >= 10:
                break
            continue
        if len(vals) <= 1 and len(all_cols) > 1:
            continue  # 섹션 헤더

        # No가 숫자인지 확인
        has_valid = False
        if tc_id_col:
            tc_val = _resolve_merged(ws, row_idx, tc_id_col)
            if tc_val and str(tc_val).strip():
                has_valid = True
        if not has_valid and no_col:
            no_val = _resolve_merged(ws, row_idx, no_col)
            if no_val is not None:
                try:
                    int(no_val)
                    has_valid = True
                except (ValueError, TypeError):
                    pass

        if has_valid:
            count += 1
            empty_streak = 0
        else:
            empty_streak += 1
            if empty_streak >= 10:
                break
    return count


def _parse_sheet(ws, project_id: int, user_id: int, db: Session, no_offset: int = 0, sheet_name: str = "기본") -> dict:
    """단일 시트를 파싱하여 TC를 DB에 추가/업데이트한다. {"created": N, "updated": N} 반환."""
    header_row = _detect_header_row(ws)
    if header_row is None:
        return 0

    col_map: dict[int, str] = {}
    for col_idx in range(1, ws.max_column + 1):
        val = ws.cell(row=header_row, column=col_idx).value
        if val and str(val).strip() in HEADER_MAP:
            mapped = HEADER_MAP[str(val).strip()]
            if mapped == "_skip":
                continue
            col_map[col_idx] = mapped

    created_count = 0
    updated_count = 0
    empty_streak = 0

    # 기존 TC 맵 (같은 프로젝트+시트+tc_id → TC 객체)
    existing_tcs = (
        db.query(TestCase)
        .filter(TestCase.project_id == project_id, TestCase.sheet_name == sheet_name, TestCase.deleted_at.is_(None))
        .all()
    )
    existing_map = {tc.tc_id: tc for tc in existing_tcs}

    for row_idx in range(header_row + 1, ws.max_row + 1):
        row_data: dict = {}
        for col_idx, field in col_map.items():
            value = _resolve_merged(ws, row_idx, col_idx)
            if value is not None:
                parsed = str(value).strip() if not isinstance(value, (int, float)) else value
                if field in row_data and row_data[field]:
                    continue
                row_data[field] = parsed

        if not row_data or all(v == "" or v is None for v in row_data.values()):
            empty_streak += 1
            if empty_streak >= 10:
                break
            continue
        empty_streak = 0

        # 섹션 헤더 행 감지
        unique_vals = set(str(v).strip() for v in row_data.values() if v and str(v).strip())
        if len(unique_vals) <= 1 and len(row_data) > 1:
            continue

        # depth3 → depth2 병합
        if row_data.get("depth3"):
            if not row_data.get("depth2"):
                row_data["depth2"] = row_data["depth3"]
            else:
                row_data["depth2"] = f"{row_data['depth2']} > {row_data['depth3']}"
        row_data.pop("depth3", None)

        if "tc_id" not in row_data or not row_data.get("tc_id"):
            if "no" in row_data:
                try:
                    no_val = int(row_data["no"])
                    row_data["tc_id"] = f"TC-{no_val:03d}"
                except (ValueError, TypeError):
                    continue
            else:
                continue

        if "no" not in row_data:
            row_data["no"] = no_offset + created_count + updated_count + 1
        else:
            try:
                row_data["no"] = int(row_data["no"]) + no_offset
            except (ValueError, TypeError):
                row_data["no"] = no_offset + created_count + updated_count + 1

        tc_id_val = str(row_data.get("tc_id", ""))
        fields = dict(
            no=row_data.get("no", no_offset + created_count + updated_count + 1),
            type=row_data.get("type"),
            category=row_data.get("category"),
            depth1=row_data.get("depth1"),
            depth2=row_data.get("depth2"),
            priority=row_data.get("priority"),
            test_type=row_data.get("test_type"),
            precondition=row_data.get("precondition"),
            test_steps=row_data.get("test_steps"),
            expected_result=row_data.get("expected_result"),
            r1=row_data.get("r1"),
            r2=row_data.get("r2"),
            r3=row_data.get("r3"),
            issue_link=row_data.get("issue_link"),
            assignee=row_data.get("assignee"),
            remarks=row_data.get("remarks"),
        )

        existing = existing_map.get(tc_id_val)
        if existing:
            # 덮어쓰기
            for key, val in fields.items():
                if val is not None:
                    setattr(existing, key, val)
            updated_count += 1
        else:
            tc = TestCase(
                project_id=project_id,
                created_by=user_id,
                tc_id=tc_id_val,
                sheet_name=sheet_name,
                **fields,
            )
            db.add(tc)
            created_count += 1

    return {"created": created_count, "updated": updated_count}


# ── Jira/Xray/Zephyr CSV 헤더 매핑 ───────────────────────────────────────────
JIRA_HEADER_MAP = {
    "Summary": "depth2",
    "Issue key": "tc_id",
    "Issue Key": "tc_id",
    "Key": "tc_id",
    "Issue id": "_skip",
    "Issue Type": "type",
    "Issue type": "type",
    "Priority": "priority",
    "Description": "test_steps",
    "Status": "_skip",
    "Component/s": "category",
    "Components": "category",
    "Component": "category",
    "Labels": "category",
    "Label": "category",
    "Assignee": "assignee",
    "Reporter": "_skip",
    "Fix Version/s": "remarks",
    "Fix Version": "remarks",
    "Affects Version/s": "_skip",
    "Created": "_skip",
    "Updated": "_skip",
    "Resolution": "_skip",
    "Epic Link": "depth1",
    "Epic Name": "depth1",
    "Sprint": "_skip",
    "Custom field (Test Steps)": "test_steps",
    "Test Step": "test_steps",
    "Test Data": "precondition",
    "Expected Result": "expected_result",
    "Precondition": "precondition",
    "Folder": "depth1",
    "Objective": "precondition",
    # Xray
    "Generic Test Definition": "test_steps",
    # Zephyr Scale
    "Name": "depth2",
    "Preconditions": "precondition",
    "Test Script (Step-by-Step) - Step": "test_steps",
    "Test Script (Step-by-Step) - Expected Result": "expected_result",
    "Test Script (Step-by-Step) - Test Data": "precondition",
}


def _parse_csv(file_content: bytes, project_id: int, user_id: int, db: Session, sheet_name: str = "CSV Import") -> dict:
    """CSV 파일을 파싱하여 TC를 DB에 추가/업데이트한다."""
    # 인코딩 감지
    text = None
    for enc in ["utf-8-sig", "utf-8", "cp949", "euc-kr", "latin-1"]:
        try:
            text = file_content.decode(enc)
            break
        except (UnicodeDecodeError, LookupError):
            continue
    if text is None:
        return {"created": 0, "updated": 0}

    reader = csv.DictReader(io.StringIO(text))
    if not reader.fieldnames:
        return {"created": 0, "updated": 0}

    # 컬럼 매핑 (HEADER_MAP + JIRA_HEADER_MAP 모두 사용)
    combined_map = {**HEADER_MAP, **JIRA_HEADER_MAP}
    col_map = {}
    for header in reader.fieldnames:
        h = header.strip()
        if h in combined_map:
            mapped = combined_map[h]
            if mapped != "_skip":
                col_map[header] = mapped

    if not col_map:
        return {"created": 0, "updated": 0}

    # 기존 TC 맵
    existing_tcs = (
        db.query(TestCase)
        .filter(TestCase.project_id == project_id, TestCase.sheet_name == sheet_name, TestCase.deleted_at.is_(None))
        .all()
    )
    existing_map = {tc.tc_id: tc for tc in existing_tcs}

    created_count = 0
    updated_count = 0

    for row_num, row in enumerate(reader, start=1):
        row_data = {}
        for csv_header, field in col_map.items():
            val = row.get(csv_header, "").strip()
            if val and field not in row_data:
                row_data[field] = val

        if not row_data or all(not v for v in row_data.values()):
            continue

        # tc_id 없으면 no 기반 생성
        if not row_data.get("tc_id"):
            if row_data.get("no"):
                try:
                    row_data["tc_id"] = f"TC-{int(row_data['no']):03d}"
                except (ValueError, TypeError):
                    row_data["tc_id"] = f"CSV-{row_num:04d}"
            else:
                row_data["tc_id"] = f"CSV-{row_num:04d}"

        if not row_data.get("no"):
            row_data["no"] = row_num

        try:
            row_data["no"] = int(row_data["no"])
        except (ValueError, TypeError):
            row_data["no"] = row_num

        # depth3 → depth2 병합
        if row_data.get("depth3"):
            if not row_data.get("depth2"):
                row_data["depth2"] = row_data["depth3"]
            else:
                row_data["depth2"] = f"{row_data['depth2']} > {row_data['depth3']}"
        row_data.pop("depth3", None)

        tc_id_val = str(row_data["tc_id"])
        fields = {k: row_data.get(k) for k in [
            "no", "type", "category", "depth1", "depth2", "priority",
            "test_type", "precondition", "test_steps", "expected_result",
            "r1", "r2", "r3", "issue_link", "assignee", "remarks",
        ] if row_data.get(k) is not None}

        existing = existing_map.get(tc_id_val)
        if existing:
            for key, val in fields.items():
                if val is not None:
                    setattr(existing, key, val)
            updated_count += 1
        else:
            tc = TestCase(
                project_id=project_id,
                created_by=user_id,
                tc_id=tc_id_val,
                sheet_name=sheet_name,
                **fields,
            )
            db.add(tc)
            created_count += 1

    return {"created": created_count, "updated": updated_count}


def _load_workbook_from_upload(file: UploadFile):
    """UploadFile을 읽어 openpyxl Workbook으로 반환한다."""
    if not file.filename or not file.filename.endswith((".xlsx", ".xls")):
        raise HTTPException(status_code=400, detail="Only .xlsx files are supported")

    file_content = file.file.read()
    if len(file_content) > MAX_IMPORT_SIZE:
        raise HTTPException(
            status_code=413,
            detail=f"File too large. Maximum size is {MAX_IMPORT_SIZE // (1024*1024)}MB",
        )

    try:
        return load_workbook(filename=io.BytesIO(file_content), data_only=True)
    except Exception:
        logger.exception("Excel read failed")
        raise HTTPException(status_code=400, detail="Failed to read Excel file")


def _is_csv_file(filename: str) -> bool:
    return filename and filename.lower().endswith(".csv")


def _is_md_file(filename: str) -> bool:
    return filename and filename.lower().endswith(".md")


# ── Markdown Import ──────────────────────────────────────────────────────────

import re

_MD_SEPARATOR_RE = re.compile(r"^\|[\s:-]+\|[\s:|-]*$")


def _decode_text(file_content: bytes) -> Optional[str]:
    """바이트를 텍스트로 디코딩한다 (CSV/MD 공용)."""
    for enc in ["utf-8-sig", "utf-8", "cp949", "euc-kr", "latin-1"]:
        try:
            return file_content.decode(enc)
        except (UnicodeDecodeError, LookupError):
            continue
    return None


def _parse_md_row(line: str) -> list[str]:
    """Markdown 테이블 행을 파싱하여 셀 리스트를 반환한다.
    이스케이프된 파이프(\\|)를 올바르게 처리한다."""
    line = line.strip()
    if line.startswith("|"):
        line = line[1:]
    if line.endswith("|"):
        line = line[:-1]
    # \| 를 플레이스홀더로 치환 후 split, 다시 복원
    placeholder = "\x00PIPE\x00"
    line = line.replace("\\|", placeholder)
    return [cell.strip().replace(placeholder, "|") for cell in line.split("|")]


def _parse_md_tables(file_content: bytes) -> list[dict]:
    """Markdown 파일에서 테이블들을 추출한다.
    반환: [{"name": str, "headers": list[str], "rows": list[list[str]]}, ...]
    """
    text = _decode_text(file_content)
    if not text:
        return []

    tables = []
    current_heading = None
    headers = None
    rows = []
    in_table = False
    unnamed_count = 0

    def _finalize_table():
        nonlocal headers, rows, in_table, unnamed_count
        if headers and rows:
            if current_heading:
                name = current_heading
            else:
                unnamed_count += 1
                name = f"MD Import" if unnamed_count == 1 else f"MD Import {unnamed_count}"
            # 중복 이름 처리
            existing_names = [t["name"] for t in tables]
            if name in existing_names:
                suffix = 2
                while f"{name} ({suffix})" in existing_names:
                    suffix += 1
                name = f"{name} ({suffix})"
            tables.append({"name": name, "headers": headers, "rows": rows})
        headers = None
        rows = []
        in_table = False

    for line in text.split("\n"):
        stripped = line.strip()

        # 헤딩 감지
        if stripped.startswith("#"):
            if in_table:
                _finalize_table()
            heading_text = stripped.lstrip("#").strip()
            if heading_text:
                current_heading = heading_text
            continue

        # 파이프 행 감지
        is_pipe_row = "|" in stripped and stripped.startswith("|")

        if is_pipe_row:
            # 구분자 행 (|---|---| 등) 스킵
            if _MD_SEPARATOR_RE.match(stripped):
                continue

            cells = _parse_md_row(stripped)

            if not in_table:
                # 첫 파이프 행 = 헤더
                headers = cells
                rows = []
                in_table = True
            else:
                rows.append(cells)
        else:
            # 테이블 밖의 줄
            if in_table:
                _finalize_table()

    # 파일 끝에 테이블이 남아있으면 저장
    if in_table:
        _finalize_table()

    return tables


def _preview_md(file_content: bytes, project_id: int, db: Session) -> list:
    """Markdown 파일의 미리보기 — 테이블별 TC 수와 기존 중복 수 반환."""
    tables = _parse_md_tables(file_content)
    result = []
    for t in tables:
        row_count = len(t["rows"])
        if row_count == 0:
            continue
        existing = db.query(TestCase).filter(
            TestCase.project_id == project_id,
            TestCase.sheet_name == t["name"],
            TestCase.deleted_at.is_(None),
        ).count()
        result.append({"name": t["name"], "tc_count": row_count, "existing": existing})
    return result


def _parse_md_table(table: dict, project_id: int, user_id: int, db: Session, sheet_name: str) -> dict:
    """파싱된 Markdown 테이블 하나를 TC로 DB에 추가/업데이트한다."""
    combined_map = {**HEADER_MAP, **JIRA_HEADER_MAP}

    # 헤더 → 필드 매핑 (인덱스 기반)
    col_map = {}
    for idx, h in enumerate(table["headers"]):
        h_stripped = h.strip()
        if h_stripped in combined_map:
            mapped = combined_map[h_stripped]
            if mapped != "_skip":
                col_map[idx] = mapped

    if not col_map:
        return {"created": 0, "updated": 0}

    # 기존 TC 맵
    existing_tcs = (
        db.query(TestCase)
        .filter(TestCase.project_id == project_id, TestCase.sheet_name == sheet_name, TestCase.deleted_at.is_(None))
        .all()
    )
    existing_map = {tc.tc_id: tc for tc in existing_tcs}

    created_count = 0
    updated_count = 0

    for row_num, cells in enumerate(table["rows"], start=1):
        row_data = {}
        for idx, field in col_map.items():
            if idx < len(cells):
                val = cells[idx].strip()
                if val and field not in row_data:
                    row_data[field] = val

        if not row_data or all(not v for v in row_data.values()):
            continue

        # tc_id 없으면 no 기반 생성
        if not row_data.get("tc_id"):
            if row_data.get("no"):
                try:
                    row_data["tc_id"] = f"TC-{int(row_data['no']):03d}"
                except (ValueError, TypeError):
                    row_data["tc_id"] = f"MD-{row_num:04d}"
            else:
                row_data["tc_id"] = f"MD-{row_num:04d}"

        if not row_data.get("no"):
            row_data["no"] = row_num

        try:
            row_data["no"] = int(row_data["no"])
        except (ValueError, TypeError):
            row_data["no"] = row_num

        # depth3 → depth2 병합
        if row_data.get("depth3"):
            if not row_data.get("depth2"):
                row_data["depth2"] = row_data["depth3"]
            else:
                row_data["depth2"] = f"{row_data['depth2']} > {row_data['depth3']}"
        row_data.pop("depth3", None)

        tc_id_val = str(row_data["tc_id"])
        fields = {k: row_data.get(k) for k in [
            "no", "type", "category", "depth1", "depth2", "priority",
            "test_type", "precondition", "test_steps", "expected_result",
            "r1", "r2", "r3", "issue_link", "assignee", "remarks",
        ] if row_data.get(k) is not None}

        existing = existing_map.get(tc_id_val)
        if existing:
            for key, val in fields.items():
                if val is not None:
                    setattr(existing, key, val)
            updated_count += 1
        else:
            tc = TestCase(
                project_id=project_id,
                created_by=user_id,
                tc_id=tc_id_val,
                sheet_name=sheet_name,
                **fields,
            )
            db.add(tc)
            created_count += 1

    return {"created": created_count, "updated": updated_count}


def _preview_csv(file_content: bytes, project_id: int, db: Session) -> list:
    """CSV 파일의 미리보기 (시트 1개로 취급)."""
    text = None
    for enc in ["utf-8-sig", "utf-8", "cp949", "euc-kr", "latin-1"]:
        try:
            text = file_content.decode(enc)
            break
        except (UnicodeDecodeError, LookupError):
            continue
    if not text:
        return []

    reader = csv.DictReader(io.StringIO(text))
    row_count = sum(1 for _ in reader)
    if row_count == 0:
        return []

    sheet_name = "CSV Import"
    existing = db.query(TestCase).filter(
        TestCase.project_id == project_id,
        TestCase.sheet_name == sheet_name,
        TestCase.deleted_at.is_(None),
    ).count()
    return [{"name": sheet_name, "tc_count": row_count, "existing": existing}]


@router.post("/import/preview")
def preview_import_sheets(
    project_id: int,
    file: UploadFile = File(...),
    db: Session = Depends(get_db),
    current_user: User = Depends(check_project_access("admin")),
):
    """엑셀/CSV/Markdown 파일의 시트 목록과 각 시트의 TC 수, 기존 중복 수를 반환한다."""
    if _is_csv_file(file.filename):
        content = file.file.read()
        if len(content) > MAX_IMPORT_SIZE:
            raise HTTPException(status_code=413, detail=f"File too large. Maximum size is {MAX_IMPORT_SIZE // (1024*1024)}MB")
        return {"sheets": _preview_csv(content, project_id, db)}

    if _is_md_file(file.filename):
        content = file.file.read()
        if len(content) > MAX_IMPORT_SIZE:
            raise HTTPException(status_code=413, detail=f"File too large. Maximum size is {MAX_IMPORT_SIZE // (1024*1024)}MB")
        return {"sheets": _preview_md(content, project_id, db)}

    wb = _load_workbook_from_upload(file)

    sheets = []
    for name in wb.sheetnames:
        ws = wb[name]
        header_row = _detect_header_row(ws)
        if header_row is None:
            continue
        tc_count = _count_tc_rows(ws, header_row)
        if tc_count == 0:
            continue
        # 기존 동일 시트의 TC 수
        existing = db.query(TestCase).filter(
            TestCase.project_id == project_id,
            TestCase.sheet_name == name,
            TestCase.deleted_at.is_(None),
        ).count()
        sheets.append({"name": name, "tc_count": tc_count, "existing": existing})

    return {"sheets": sheets}


@router.post("/import", status_code=status.HTTP_201_CREATED)
def import_testcases(
    project_id: int,
    file: UploadFile = File(...),
    sheet_names: Optional[str] = Query(None, description="쉼표 구분 시트명 (미지정 시 전체)"),
    db: Session = Depends(get_db),
    current_user: User = Depends(check_project_access("admin")),
):
    _get_project_or_404(project_id, db)

    # CSV 파일 처리
    if _is_csv_file(file.filename):
        content = file.file.read()
        if len(content) > MAX_IMPORT_SIZE:
            raise HTTPException(status_code=413, detail=f"File too large. Maximum size is {MAX_IMPORT_SIZE // (1024*1024)}MB")
        sheet_name = "CSV Import"
        from models import TestCaseSheet
        sheet_exists = db.query(TestCaseSheet).filter(
            TestCaseSheet.project_id == project_id, TestCaseSheet.name == sheet_name
        ).first()
        if not sheet_exists:
            max_order = db.query(TestCaseSheet.sort_order).filter(
                TestCaseSheet.project_id == project_id
            ).order_by(TestCaseSheet.sort_order.desc()).first()
            db.add(TestCaseSheet(project_id=project_id, name=sheet_name, sort_order=(max_order[0] + 1) if max_order else 0))
            db.flush()
        r = _parse_csv(content, project_id, current_user.id, db, sheet_name=sheet_name)
        db.commit()
        return {"created": r["created"], "updated": r["updated"], "imported": r["created"] + r["updated"], "sheets": [{"sheet": sheet_name, "created": r["created"], "updated": r["updated"]}]}

    # Markdown 파일 처리
    if _is_md_file(file.filename):
        content = file.file.read()
        if len(content) > MAX_IMPORT_SIZE:
            raise HTTPException(status_code=413, detail=f"File too large. Maximum size is {MAX_IMPORT_SIZE // (1024*1024)}MB")
        tables = _parse_md_tables(content)
        if not tables:
            return {"created": 0, "updated": 0, "imported": 0, "sheets": []}

        if sheet_names:
            target_names = [s.strip() for s in sheet_names.split(",") if s.strip()]
        else:
            target_names = [t["name"] for t in tables]

        from models import TestCaseSheet
        total_created = 0
        total_updated = 0
        results = []
        for table in tables:
            if table["name"] not in target_names:
                continue
            # 시트 레코드 자동 생성
            sheet_exists = db.query(TestCaseSheet).filter(
                TestCaseSheet.project_id == project_id, TestCaseSheet.name == table["name"]
            ).first()
            if not sheet_exists:
                max_order = db.query(TestCaseSheet.sort_order).filter(
                    TestCaseSheet.project_id == project_id
                ).order_by(TestCaseSheet.sort_order.desc()).first()
                db.add(TestCaseSheet(project_id=project_id, name=table["name"], sort_order=(max_order[0] + 1) if max_order else 0))
                db.flush()
            r = _parse_md_table(table, project_id, current_user.id, db, sheet_name=table["name"])
            results.append({"sheet": table["name"], "created": r["created"], "updated": r["updated"]})
            total_created += r["created"]
            total_updated += r["updated"]

        db.commit()
        return {"created": total_created, "updated": total_updated, "imported": total_created + total_updated, "sheets": results}

    wb = _load_workbook_from_upload(file)

    # 대상 시트 결정
    if sheet_names:
        target_names = [s.strip() for s in sheet_names.split(",") if s.strip()]
    else:
        # 미지정 시: 유효한 시트 전부 (SKIP_SHEETS 제외)
        target_names = []
        for name in wb.sheetnames:
            if name.upper() in SKIP_SHEETS:
                continue
            ws = wb[name]
            if _detect_header_row(ws) is not None:
                target_names.append(name)
        # 아무것도 없으면 첫 시트
        if not target_names and wb.sheetnames:
            target_names = [wb.sheetnames[0]]

    total_created = 0
    total_updated = 0
    results = []
    for name in target_names:
        if name not in wb.sheetnames:
            continue
        ws = wb[name]
        # 시트 레코드 자동 생성
        from models import TestCaseSheet
        sheet_exists = db.query(TestCaseSheet).filter(
            TestCaseSheet.project_id == project_id, TestCaseSheet.name == name
        ).first()
        if not sheet_exists:
            max_order = db.query(TestCaseSheet.sort_order).filter(
                TestCaseSheet.project_id == project_id
            ).order_by(TestCaseSheet.sort_order.desc()).first()
            db.add(TestCaseSheet(project_id=project_id, name=name, sort_order=(max_order[0] + 1) if max_order else 0))
            db.flush()

        r = _parse_sheet(ws, project_id, current_user.id, db, no_offset=0, sheet_name=name)
        results.append({"sheet": name, "created": r["created"], "updated": r["updated"]})
        total_created += r["created"]
        total_updated += r["updated"]

    db.commit()
    return {"created": total_created, "updated": total_updated, "imported": total_created + total_updated, "sheets": results}


# ── Export Excel ──────────────────────────────────────────────────────────────

@router.get("/export")
def export_testcases(
    project_id: int,
    db: Session = Depends(get_db),
    current_user: User = Depends(check_project_access("viewer")),
):
    project = _get_project_or_404(project_id, db)
    testcases = (
        db.query(TestCase)
        .filter(TestCase.project_id == project_id)
        .order_by(TestCase.no)
        .all()
    )

    wb = Workbook()
    ws = wb.active
    ws.title = "Test Cases"

    # -- Styles --
    dark_fill = PatternFill(start_color="2F3136", end_color="2F3136", fill_type="solid")
    header_font = Font(name="Malgun Gothic", bold=True, color="FFFFFF", size=10)
    cell_font = Font(name="Malgun Gothic", size=10)
    thin_border = Border(
        left=Side(style="thin"),
        right=Side(style="thin"),
        top=Side(style="thin"),
        bottom=Side(style="thin"),
    )
    center_align = Alignment(horizontal="center", vertical="center", wrap_text=True)
    left_align = Alignment(horizontal="left", vertical="center", wrap_text=True)

    pass_fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
    fail_fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
    block_fill = PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid")
    na_fill = PatternFill(start_color="D9D9D9", end_color="D9D9D9", fill_type="solid")

    # Title rows (rows 1-4)
    ws.merge_cells("B1:S1")
    title_cell = ws["B1"]
    title_cell.value = f"{project.name} - Test Cases"
    title_cell.font = Font(name="Malgun Gothic", bold=True, size=14)
    title_cell.alignment = Alignment(horizontal="center", vertical="center")

    ws.merge_cells("B3:S3")
    ws["B3"].value = f"Project: {project.name}"
    ws["B3"].font = Font(name="Malgun Gothic", size=10)

    # Headers in row 5 (columns B through S = columns 2-19)
    headers = [
        "No", "TC ID", "Type", "Category", "Depth1", "Depth2",
        "Priority", "Platform", "Precondition", "Steps",
        "Expected Result", "Issue Link", "Assignee", "Remarks",
        "Result", "Actual Result", "Issue Link (Run)", "Run Remarks",
    ]

    col_widths = [6, 10, 10, 15, 18, 18, 10, 12, 25, 35, 35, 20, 12, 20, 10, 35, 20, 20]

    for idx, (header, width) in enumerate(zip(headers, col_widths)):
        col = idx + 2  # start at column B
        cell = ws.cell(row=5, column=col, value=header)
        cell.font = header_font
        cell.fill = dark_fill
        cell.alignment = center_align
        cell.border = thin_border
        ws.column_dimensions[get_column_letter(col)].width = width

    # Data rows starting at row 6
    for row_offset, tc in enumerate(testcases):
        row = 6 + row_offset
        values = [
            tc.no, tc.tc_id, tc.type, tc.category, tc.depth1, tc.depth2,
            tc.priority, tc.test_type, tc.precondition, tc.test_steps,
            tc.expected_result, tc.issue_link, tc.assignee, tc.remarks,
            "", "", "", "",
        ]
        for col_offset, value in enumerate(values):
            col = col_offset + 2
            cell = ws.cell(row=row, column=col, value=value)
            cell.font = cell_font
            cell.border = thin_border
            if col_offset < 2 or col_offset == 6 or col_offset == 7 or col_offset == 12 or col_offset == 14:
                cell.alignment = center_align
            else:
                cell.alignment = left_align

    # Apply conditional fill for result column (column index 16 = col P)
    result_col = 16  # column P (0-based offset 14 + 2)
    for row in range(6, 6 + len(testcases)):
        cell = ws.cell(row=row, column=result_col)
        val = str(cell.value or "").upper()
        if val == "PASS":
            cell.fill = pass_fill
        elif val == "FAIL":
            cell.fill = fail_fill
        elif val == "BLOCK":
            cell.fill = block_fill
        elif val in ("NA", "N/A"):
            cell.fill = na_fill

    # Freeze pane below header
    ws.freeze_panes = "B6"

    output = io.BytesIO()
    wb.save(output)
    output.seek(0)

    from urllib.parse import quote
    filename = f"{project.name}_TestCases.xlsx"
    encoded = quote(filename)
    return StreamingResponse(
        output,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename*=UTF-8''{encoded}"},
    )

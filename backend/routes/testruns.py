import io
import os
from datetime import datetime
from models import now_kst
from typing import List

from fastapi import APIRouter, Depends, HTTPException, status, Query
from fastapi.responses import StreamingResponse
from sqlalchemy.orm import Session, joinedload
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter

from database import get_db
from models import User, Project, TestCase, TestRun, TestResult, TestRunStatus, TestResultValue, Attachment
from schemas import (
    TestRunCreate, TestRunUpdate, TestRunResponse, TestRunListResponse,
    TestResultCreate, TestResultResponse,
)
from auth import get_current_user, role_required, check_project_access
from routes.attachments import UPLOAD_DIR

router = APIRouter(
    prefix="/api/projects/{project_id}/testruns",
    tags=["testruns"],
)


def _get_project_or_404(project_id: int, db: Session) -> Project:
    project = db.query(Project).filter(Project.id == project_id).first()
    if not project:
        raise HTTPException(status_code=404, detail="Project not found")
    return project


@router.get("", response_model=List[TestRunListResponse])
def list_testruns(
    project_id: int,
    limit: int = Query(500, ge=1, le=1000),
    offset: int = Query(0, ge=0),
    db: Session = Depends(get_db),
    current_user: User = Depends(check_project_access("viewer")),
):
    _get_project_or_404(project_id, db)
    return (
        db.query(TestRun)
        .filter(TestRun.project_id == project_id)
        .order_by(TestRun.created_at.desc())
        .offset(offset)
        .limit(limit)
        .all()
    )


@router.post("", response_model=TestRunListResponse, status_code=status.HTTP_201_CREATED)
def create_testrun(
    project_id: int,
    payload: TestRunCreate,
    db: Session = Depends(get_db),
    current_user: User = Depends(check_project_access("tester")),
):
    _get_project_or_404(project_id, db)

    # test_plan_id가 지정된 경우 같은 프로젝트의 플랜인지 검증
    if payload.test_plan_id is not None:
        from models import TestPlan
        plan = db.query(TestPlan).filter(TestPlan.id == payload.test_plan_id).first()
        if not plan:
            raise HTTPException(status_code=400, detail="존재하지 않는 테스트 플랜입니다.")
        if plan.project_id != project_id:
            raise HTTPException(status_code=400, detail="다른 프로젝트의 테스트 플랜은 연결할 수 없습니다.")

    run = TestRun(
        project_id=project_id,
        name=payload.name,
        version=payload.version,
        environment=payload.environment,
        round=payload.round,
        test_plan_id=payload.test_plan_id,
        created_by=current_user.id,
    )
    db.add(run)
    db.flush()  # get run.id

    # Auto-create empty TestResult for each TC in the project
    tcs = db.query(TestCase).filter(TestCase.project_id == project_id, TestCase.deleted_at.is_(None)).order_by(TestCase.no).all()
    for tc in tcs:
        tr = TestResult(
            test_run_id=run.id,
            test_case_id=tc.id,
            result=TestResultValue.NS,
            executed_by=current_user.id,
        )
        db.add(tr)

    db.commit()
    db.refresh(run)
    return run


@router.get("/{run_id}", response_model=TestRunResponse)
def get_testrun(
    project_id: int,
    run_id: int,
    db: Session = Depends(get_db),
    current_user: User = Depends(check_project_access("viewer")),
):
    _get_project_or_404(project_id, db)

    run = (
        db.query(TestRun)
        .options(joinedload(TestRun.results).joinedload(TestResult.test_case))
        .filter(TestRun.id == run_id, TestRun.project_id == project_id)
        .first()
    )
    if not run:
        raise HTTPException(status_code=404, detail="Test run not found")
    return run


@router.put("/{run_id}", response_model=TestRunListResponse)
def update_testrun(
    project_id: int,
    run_id: int,
    payload: TestRunUpdate,
    db: Session = Depends(get_db),
    current_user: User = Depends(check_project_access("tester")),
):
    _get_project_or_404(project_id, db)

    run = db.query(TestRun).filter(
        TestRun.id == run_id, TestRun.project_id == project_id
    ).first()
    if not run:
        raise HTTPException(status_code=404, detail="Test run not found")

    for key, value in payload.model_dump(exclude_unset=True).items():
        setattr(run, key, value)

    db.commit()
    db.refresh(run)
    return run


@router.post("/{run_id}/results", response_model=List[TestResultResponse])
def submit_results(
    project_id: int,
    run_id: int,
    results: List[TestResultCreate],
    db: Session = Depends(get_db),
    current_user: User = Depends(check_project_access("tester")),
):
    _get_project_or_404(project_id, db)

    run = db.query(TestRun).filter(
        TestRun.id == run_id, TestRun.project_id == project_id
    ).first()
    if not run:
        raise HTTPException(status_code=404, detail="Test run not found")

    if run.status == TestRunStatus.completed:
        raise HTTPException(status_code=400, detail="완료된 테스트 런은 수정할 수 없습니다. 재오픈 후 수정하세요.")

    # Validate all test_case_ids belong to this project
    tc_ids = [r.test_case_id for r in results]
    valid_tc_ids = set(
        row[0] for row in db.query(TestCase.id).filter(
            TestCase.project_id == project_id, TestCase.id.in_(tc_ids)
        ).all()
    )
    invalid_ids = set(tc_ids) - valid_tc_ids
    if invalid_ids:
        raise HTTPException(
            status_code=400,
            detail="One or more test case IDs are invalid for this project",
        )

    saved: list[TestResult] = []
    for r in results:
        # Validate result enum
        try:
            result_enum = TestResultValue(r.result)
        except ValueError:
            raise HTTPException(
                status_code=400,
                detail=f"Invalid result value: {r.result}. Must be one of PASS, FAIL, BLOCK, NA, NS",
            )

        # Check if result already exists for this run + test case
        existing = db.query(TestResult).filter(
            TestResult.test_run_id == run_id,
            TestResult.test_case_id == r.test_case_id,
        ).first()

        # Parse optional time fields
        started = None
        finished = None
        if r.started_at:
            try:
                started = datetime.fromisoformat(r.started_at)
            except (ValueError, TypeError):
                pass
        if r.finished_at:
            try:
                finished = datetime.fromisoformat(r.finished_at)
            except (ValueError, TypeError):
                pass

        if existing:
            existing.result = result_enum
            existing.actual_result = r.actual_result
            existing.issue_link = r.issue_link
            existing.remarks = r.remarks
            existing.executed_by = current_user.id
            existing.executed_at = now_kst()
            if started:
                existing.started_at = started
            if finished:
                existing.finished_at = finished
            if r.duration_sec is not None:
                existing.duration_sec = r.duration_sec
            saved.append(existing)
        else:
            tr = TestResult(
                test_run_id=run_id,
                test_case_id=r.test_case_id,
                result=result_enum,
                actual_result=r.actual_result,
                issue_link=r.issue_link,
                remarks=r.remarks,
                executed_by=current_user.id,
            )
            db.add(tr)
            saved.append(tr)

    db.commit()
    for item in saved:
        db.refresh(item)
    return saved


@router.put("/{run_id}/complete", response_model=TestRunListResponse)
def complete_testrun(
    project_id: int,
    run_id: int,
    db: Session = Depends(get_db),
    current_user: User = Depends(check_project_access("tester")),
):
    _get_project_or_404(project_id, db)

    run = db.query(TestRun).filter(
        TestRun.id == run_id, TestRun.project_id == project_id
    ).first()
    if not run:
        raise HTTPException(status_code=404, detail="Test run not found")

    run.status = TestRunStatus.completed
    run.completed_at = now_kst()

    db.commit()
    db.refresh(run)
    return run


@router.put("/{run_id}/reopen", response_model=TestRunListResponse)
def reopen_testrun(
    project_id: int,
    run_id: int,
    db: Session = Depends(get_db),
    current_user: User = Depends(check_project_access("tester")),
):
    """완료된 테스트 런을 다시 진행 중으로 변경한다."""
    _get_project_or_404(project_id, db)

    run = db.query(TestRun).filter(
        TestRun.id == run_id, TestRun.project_id == project_id
    ).first()
    if not run:
        raise HTTPException(status_code=404, detail="Test run not found")

    run.status = TestRunStatus.in_progress
    run.completed_at = None
    db.commit()
    db.refresh(run)
    return run


@router.delete("/{run_id}", status_code=status.HTTP_204_NO_CONTENT)
def delete_testrun(
    project_id: int,
    run_id: int,
    db: Session = Depends(get_db),
    current_user: User = Depends(check_project_access("admin")),
):
    _get_project_or_404(project_id, db)

    run = db.query(TestRun).filter(
        TestRun.id == run_id, TestRun.project_id == project_id
    ).first()
    if not run:
        raise HTTPException(status_code=404, detail="Test run not found")

    result_ids = [r.id for r in db.query(TestResult.id).filter(TestResult.test_run_id == run_id).all()]
    if result_ids:
        attachments = db.query(Attachment).filter(Attachment.test_result_id.in_(result_ids)).all()
        for att in attachments:
            if att.filepath:
                full_path = os.path.join(UPLOAD_DIR, att.filepath)
                if os.path.isfile(full_path):
                    try:
                        os.remove(full_path)
                    except OSError:
                        pass
        # bulk delete로 처리 (개별 db.delete()는 cascade와 충돌하여 경고 발생)
        db.query(Attachment).filter(Attachment.test_result_id.in_(result_ids)).delete(synchronize_session=False)
    db.query(TestResult).filter(TestResult.test_run_id == run_id).delete(synchronize_session=False)
    db.delete(run)
    db.commit()


@router.post("/{run_id}/clone", response_model=TestRunListResponse, status_code=status.HTTP_201_CREATED)
def clone_testrun(
    project_id: int,
    run_id: int,
    db: Session = Depends(get_db),
    current_user: User = Depends(check_project_access("tester")),
):
    """Clone an existing test run with all its test results reset to NS."""
    _get_project_or_404(project_id, db)

    source = (
        db.query(TestRun)
        .options(joinedload(TestRun.results))
        .filter(TestRun.id == run_id, TestRun.project_id == project_id)
        .first()
    )
    if not source:
        raise HTTPException(status_code=404, detail="Test run not found")

    new_run = TestRun(
        project_id=project_id,
        name=f"{source.name} (복제)",
        version=source.version,
        environment=source.environment,
        round=source.round,
        created_by=current_user.id,
    )
    db.add(new_run)
    db.flush()

    for r in source.results:
        tr = TestResult(
            test_run_id=new_run.id,
            test_case_id=r.test_case_id,
            result=TestResultValue.NS,
            executed_by=current_user.id,
        )
        db.add(tr)

    db.commit()
    db.refresh(new_run)
    return new_run


@router.get("/{run_id}/export")
def export_testrun_excel(
    project_id: int,
    run_id: int,
    db: Session = Depends(get_db),
    current_user: User = Depends(check_project_access("viewer")),
):
    """Export test run results as Excel file."""
    _get_project_or_404(project_id, db)

    run = (
        db.query(TestRun)
        .options(joinedload(TestRun.results).joinedload(TestResult.test_case))
        .filter(TestRun.id == run_id, TestRun.project_id == project_id)
        .first()
    )
    if not run:
        raise HTTPException(status_code=404, detail="Test run not found")

    wb = Workbook()
    ws = wb.active
    ws.title = "Test Results"

    # Header style
    header_font = Font(bold=True, color="FFFFFF", size=10)
    header_fill = PatternFill(start_color="1A2744", end_color="1A2744", fill_type="solid")
    header_align = Alignment(horizontal="center", vertical="center", wrap_text=True)

    headers = ["No", "TC ID", "Type", "Category", "Depth1", "Depth2", "Priority",
               "Test Steps", "Expected Result", "Result", "Actual Result",
               "Issue Link", "Duration(sec)", "Remarks"]
    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_align

    # Result color fills
    result_fills = {
        "PASS": PatternFill(start_color="DCFCE7", end_color="DCFCE7", fill_type="solid"),
        "FAIL": PatternFill(start_color="FEE2E2", end_color="FEE2E2", fill_type="solid"),
        "BLOCK": PatternFill(start_color="FEF9C3", end_color="FEF9C3", fill_type="solid"),
        "NA": PatternFill(start_color="F3F4F6", end_color="F3F4F6", fill_type="solid"),
    }

    # Sort results by test_case.no
    sorted_results = sorted(run.results, key=lambda r: (r.test_case.no if r.test_case else 0))

    for row_idx, tr in enumerate(sorted_results, 2):
        tc = tr.test_case
        result_display = "" if tr.result == TestResultValue.NS else ("N/A" if tr.result == TestResultValue.NA else tr.result.value)

        ws.cell(row=row_idx, column=1, value=tc.no if tc else "")
        ws.cell(row=row_idx, column=2, value=tc.tc_id if tc else "")
        ws.cell(row=row_idx, column=3, value=tc.type if tc else "")
        ws.cell(row=row_idx, column=4, value=tc.category if tc else "")
        ws.cell(row=row_idx, column=5, value=tc.depth1 if tc else "")
        ws.cell(row=row_idx, column=6, value=tc.depth2 if tc else "")
        ws.cell(row=row_idx, column=7, value=tc.priority if tc else "")
        ws.cell(row=row_idx, column=8, value=tc.test_steps if tc else "").alignment = Alignment(wrap_text=True)
        ws.cell(row=row_idx, column=9, value=tc.expected_result if tc else "").alignment = Alignment(wrap_text=True)
        result_cell = ws.cell(row=row_idx, column=10, value=result_display)
        result_cell.alignment = Alignment(horizontal="center")
        if tr.result.value in result_fills:
            result_cell.fill = result_fills[tr.result.value]
        ws.cell(row=row_idx, column=11, value=tr.actual_result or "").alignment = Alignment(wrap_text=True)
        ws.cell(row=row_idx, column=12, value=tr.issue_link or "")
        ws.cell(row=row_idx, column=13, value=tr.duration_sec or "")
        ws.cell(row=row_idx, column=14, value=tr.remarks or "").alignment = Alignment(wrap_text=True)

    # Auto-width (approximate)
    col_widths = [6, 12, 8, 14, 14, 14, 10, 40, 30, 10, 30, 20, 10, 20]
    for i, w in enumerate(col_widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)

    filename = f"{run.name}_results.xlsx"
    return StreamingResponse(
        buf,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f'attachment; filename="{filename}"'},
    )

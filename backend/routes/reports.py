import io
import os
from typing import List

from fastapi import APIRouter, Depends, HTTPException
from fastapi.responses import StreamingResponse
from sqlalchemy.orm import Session, joinedload, load_only
from sqlalchemy import func, case
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

import logging

from database import get_db
from models import (
    User, Project, TestRun, TestResult, TestCase, TestResultValue,
)
from auth import get_current_user, check_project_access

logger = logging.getLogger(__name__)

router = APIRouter(
    prefix="/api/projects/{project_id}/reports",
    tags=["reports"],
)


def _get_project_or_404(project_id: int, db: Session) -> Project:
    project = db.query(Project).filter(Project.id == project_id).first()
    if not project:
        raise HTTPException(status_code=404, detail="Project not found")
    return project


def _get_run_or_404(project_id: int, test_run_id: int, db: Session) -> TestRun:
    run = db.query(TestRun).filter(
        TestRun.id == test_run_id, TestRun.project_id == project_id
    ).first()
    if not run:
        raise HTTPException(status_code=404, detail="Test run not found")
    return run


def _summary_sql(run_id: int, db: Session) -> dict:
    """SQL 집계로 summary 생성 (Python 루프 대신)."""
    row = db.query(
        func.count(TestResult.id).label("total"),
        func.sum(case((TestResult.result == TestResultValue.PASS, 1), else_=0)).label("passed"),
        func.sum(case((TestResult.result == TestResultValue.FAIL, 1), else_=0)).label("failed"),
        func.sum(case((TestResult.result == TestResultValue.BLOCK, 1), else_=0)).label("blocked"),
        func.sum(case((TestResult.result == TestResultValue.NA, 1), else_=0)).label("na"),
        func.sum(case((TestResult.result == TestResultValue.NS, 1), else_=0)).label("ns"),
    ).filter(TestResult.test_run_id == run_id).first()

    total = row.total or 0
    passed = row.passed or 0
    failed = row.failed or 0
    blocked = row.blocked or 0
    na = row.na or 0
    ns = row.ns or 0
    executed = passed + failed + blocked
    pass_rate = round(passed / executed * 100, 1) if executed > 0 else 0.0

    return {
        "total": total, "passed": passed, "failed": failed,
        "blocked": blocked, "na": na, "ns": ns, "pass_rate": pass_rate,
    }


def _category_summary_sql(run_id: int, db: Session) -> list:
    """SQL 집계로 카테고리별 통계 생성."""
    rows = (
        db.query(
            TestCase.category,
            func.count(TestResult.id).label("total"),
            func.sum(case((TestResult.result == TestResultValue.PASS, 1), else_=0)).label("passed"),
            func.sum(case((TestResult.result == TestResultValue.FAIL, 1), else_=0)).label("failed"),
            func.sum(case((TestResult.result == TestResultValue.BLOCK, 1), else_=0)).label("blocked"),
            func.sum(case((TestResult.result == TestResultValue.NA, 1), else_=0)).label("na"),
            func.sum(case((TestResult.result == TestResultValue.NS, 1), else_=0)).label("ns"),
        )
        .join(TestResult, TestResult.test_case_id == TestCase.id)
        .filter(TestResult.test_run_id == run_id)
        .group_by(TestCase.category)
        .all()
    )
    return [
        {
            "category": r.category or "Uncategorized",
            "total": r.total or 0,
            "passed": r.passed or 0,
            "failed": r.failed or 0,
            "blocked": r.blocked or 0,
            "na": r.na or 0,
            "not_started": r.ns or 0,
        }
        for r in rows
    ]


def _failed_items(run_id: int, db: Session) -> list:
    """FAIL 결과만 필요한 컬럼으로 조회."""
    results = (
        db.query(TestResult)
        .options(
            joinedload(TestResult.test_case).load_only(
                TestCase.tc_id, TestCase.category, TestCase.depth1,
                TestCase.depth2, TestCase.test_steps, TestCase.expected_result,
            )
        )
        .filter(
            TestResult.test_run_id == run_id,
            TestResult.result == TestResultValue.FAIL,
        )
        .all()
    )
    return [
        {
            "tc_id": r.test_case.tc_id,
            "category": r.test_case.category,
            "depth1": r.test_case.depth1,
            "depth2": r.test_case.depth2,
            "test_steps": r.test_case.test_steps,
            "expected_result": r.test_case.expected_result,
            "actual_result": r.actual_result,
            "issue_link": r.issue_link,
        }
        for r in results
    ]


def _build_report_data(run: TestRun, db: Session) -> dict:
    """SQL 집계 기반 리포트 데이터 생성 (전체 ORM 로드 없음)."""
    summary = _summary_sql(run.id, db)
    categories = _category_summary_sql(run.id, db)
    failed = _failed_items(run.id, db)

    return {
        "run": {
            "id": run.id,
            "name": run.name,
            "version": run.version,
            "environment": run.environment,
            "round": run.round,
            "status": run.status.value if hasattr(run.status, "value") else run.status,
            "created_at": run.created_at.isoformat() if run.created_at else None,
            "completed_at": run.completed_at.isoformat() if run.completed_at else None,
        },
        "summary": summary,
        "categories": categories,
        "failed_items": failed,
    }


# ── JSON report ───────────────────────────────────────────────────────────────

@router.get("")
def report_json(
    project_id: int,
    run_id: int = None,
    db: Session = Depends(get_db),
    current_user: User = Depends(check_project_access("viewer")),
):
    if not run_id:
        raise HTTPException(status_code=400, detail="run_id is required")
    project = _get_project_or_404(project_id, db)
    run = _get_run_or_404(project_id, run_id, db)
    raw = _build_report_data(run, db)

    # Map to frontend ReportData format
    summary = raw["summary"]
    return {
        "project": {
            "id": project.id,
            "name": project.name,
            "description": project.description,
            "jira_base_url": project.jira_base_url,
            "created_by": project.created_by,
            "created_at": project.created_at.isoformat() if project.created_at else None,
            "updated_at": project.updated_at.isoformat() if project.updated_at else None,
        },
        "test_run": raw["run"],
        "summary": {
            "total": summary["total"],
            "pass": summary["passed"],
            "fail": summary["failed"],
            "block": summary["blocked"],
            "na": summary["na"],
            "not_started": summary["ns"],
            "pass_rate": summary["pass_rate"],
            "fail_rate": round(summary["failed"] / summary["total"] * 100, 1) if summary["total"] > 0 else 0.0,
            "block_rate": round(summary["blocked"] / summary["total"] * 100, 1) if summary["total"] > 0 else 0.0,
            "na_rate": round(summary["na"] / summary["total"] * 100, 1) if summary["total"] > 0 else 0.0,
            "not_started_rate": round(summary["ns"] / summary["total"] * 100, 1) if summary["total"] > 0 else 0.0,
        },
        "top_failures": [
            {
                "test_case": {"tc_id": f["tc_id"]},
                "result": "FAIL",
                "actual_result": f.get("actual_result"),
                "issue_link": f.get("issue_link"),
            }
            for f in raw["failed_items"]
        ],
        "jira_issues": list({
            f["issue_link"]
            for f in raw["failed_items"]
            if f.get("issue_link")
        }),
        "category_summary": [
            {
                "category": c["category"],
                "total": c["total"],
                "pass": c["passed"],
                "fail": c["failed"],
                "block": c["blocked"],
                "na": c.get("na", 0),
                "not_started": c.get("not_started", 0),
            }
            for c in raw["categories"]
        ],
    }


# ── PDF report ────────────────────────────────────────────────────────────────

@router.get("/pdf")
def report_pdf(
    project_id: int,
    run_id: int = None,
    db: Session = Depends(get_db),
    current_user: User = Depends(check_project_access("viewer")),
):
    from fpdf import FPDF

    if not run_id:
        raise HTTPException(status_code=400, detail="run_id is required")
    project = _get_project_or_404(project_id, db)
    run = _get_run_or_404(project_id, run_id, db)
    data = _build_report_data(run, db)

    pdf = FPDF()
    pdf.set_auto_page_break(auto=True, margin=15)
    pdf.add_page()

    # Try to load Korean font for proper rendering
    font_name = "Helvetica"
    korean_font_paths = [
        # 맑은 고딕 (Windows)
        "C:/Windows/Fonts/malgun.ttf",
        # NanumGothic (Linux)
        "/usr/share/fonts/truetype/nanum/NanumGothic.ttf",
        "/usr/share/fonts/truetype/malgun.ttf",
        # 프로젝트 내장 폰트
        os.path.join(os.path.dirname(__file__), "..", "fonts", "malgun.ttf"),
    ]
    font_loaded = False
    for font_path in korean_font_paths:
        if os.path.exists(font_path):
            pdf.add_font("MalgunGothic", "", font_path, uni=True)
            pdf.add_font("MalgunGothic", "B", font_path, uni=True)
            font_name = "MalgunGothic"
            font_loaded = True
            break

    if not font_loaded:
        logger.warning(
            "한글 폰트를 찾을 수 없습니다. PDF에 한글이 깨질 수 있습니다. "
            "탐색 경로: %s", korean_font_paths
        )

    # Title
    pdf.set_font(font_name, "B", 16)
    pdf.cell(0, 12, f"{project.name} - Test Report", new_x="LMARGIN", new_y="NEXT", align="C")
    pdf.ln(4)

    # Run info
    pdf.set_font(font_name, "", 10)
    run_info = data["run"]
    pdf.cell(0, 7, f"Test Run: {run_info['name']}", new_x="LMARGIN", new_y="NEXT")
    pdf.cell(0, 7, f"Version: {run_info.get('version', 'N/A')}  |  Environment: {run_info.get('environment', 'N/A')}  |  Round: {run_info['round']}", new_x="LMARGIN", new_y="NEXT")
    pdf.cell(0, 7, f"Status: {run_info['status']}  |  Created: {run_info.get('created_at', 'N/A')}", new_x="LMARGIN", new_y="NEXT")
    pdf.ln(6)

    # Summary
    pdf.set_font(font_name, "B", 13)
    pdf.cell(0, 10, "Summary", new_x="LMARGIN", new_y="NEXT")
    pdf.set_font(font_name, "", 10)

    summary = data["summary"]
    pdf.set_fill_color(240, 240, 240)

    col_w = 27
    headers = ["Total", "Pass", "Fail", "Block", "NA", "NS", "Pass Rate"]
    values = [
        str(summary["total"]), str(summary["passed"]), str(summary["failed"]),
        str(summary["blocked"]), str(summary["na"]), str(summary["ns"]),
        f"{summary['pass_rate']}%",
    ]

    pdf.set_font(font_name, "B", 9)
    for h in headers:
        pdf.cell(col_w, 8, h, border=1, align="C", fill=True)
    pdf.ln()

    pdf.set_font(font_name, "", 9)
    for v in values:
        pdf.cell(col_w, 8, v, border=1, align="C")
    pdf.ln(12)

    # Category breakdown
    if data["categories"]:
        pdf.set_font(font_name, "B", 13)
        pdf.cell(0, 10, "Category Breakdown", new_x="LMARGIN", new_y="NEXT")

        cat_headers = ["Category", "Total", "Pass", "Fail", "Block"]
        cat_widths = [60, 25, 25, 25, 25]

        pdf.set_font(font_name, "B", 9)
        for h, w in zip(cat_headers, cat_widths):
            pdf.cell(w, 8, h, border=1, align="C", fill=True)
        pdf.ln()

        pdf.set_font(font_name, "", 9)
        for cat in data["categories"]:
            vals = [
                str(cat["category"])[:30],
                str(cat["total"]),
                str(cat["passed"]),
                str(cat["failed"]),
                str(cat["blocked"]),
            ]
            for v, w in zip(vals, cat_widths):
                pdf.cell(w, 8, v, border=1, align="C")
            pdf.ln()
        pdf.ln(8)

    # Failed items
    if data["failed_items"]:
        pdf.set_font(font_name, "B", 13)
        pdf.cell(0, 10, "Failed Test Cases", new_x="LMARGIN", new_y="NEXT")

        pdf.set_font(font_name, "", 9)
        for idx, item in enumerate(data["failed_items"], 1):
            pdf.set_font(font_name, "B", 9)
            pdf.cell(0, 7, f"{idx}. {item['tc_id']} - {item.get('depth1', '')} / {item.get('depth2', '')}", new_x="LMARGIN", new_y="NEXT")
            pdf.set_font(font_name, "", 8)
            if item.get("test_steps"):
                steps_text = str(item["test_steps"])[:200]
                pdf.multi_cell(0, 5, f"   Steps: {steps_text}", new_x="LMARGIN", new_y="NEXT")
            if item.get("expected_result"):
                pdf.multi_cell(0, 5, f"   Expected: {str(item['expected_result'])[:200]}", new_x="LMARGIN", new_y="NEXT")
            if item.get("actual_result"):
                pdf.multi_cell(0, 5, f"   Actual: {str(item['actual_result'])[:200]}", new_x="LMARGIN", new_y="NEXT")
            if item.get("issue_link"):
                pdf.cell(0, 5, f"   Issue: {item['issue_link']}", new_x="LMARGIN", new_y="NEXT")
            pdf.ln(3)

    output = io.BytesIO()
    pdf.output(output)
    output.seek(0)

    from urllib.parse import quote
    filename = f"{project.name}_Report_R{run.round}.pdf"
    encoded = quote(filename)
    return StreamingResponse(
        output,
        media_type="application/pdf",
        headers={"Content-Disposition": f"attachment; filename*=UTF-8''{encoded}"},
    )


# ── Excel report ──────────────────────────────────────────────────────────────

@router.get("/excel")
def report_excel(
    project_id: int,
    run_id: int = None,
    db: Session = Depends(get_db),
    current_user: User = Depends(check_project_access("viewer")),
):
    if not run_id:
        raise HTTPException(status_code=400, detail="run_id is required")
    project = _get_project_or_404(project_id, db)
    run = _get_run_or_404(project_id, run_id, db)

    # SQL 집계로 summary + categories (전체 ORM 로드 없음)
    summary = _summary_sql(run.id, db)

    # Excel 결과 시트용: 1회만 조회 (이중 로드 제거), 필요한 컬럼만 load
    results = (
        db.query(TestResult)
        .options(
            joinedload(TestResult.test_case).load_only(
                TestCase.no, TestCase.tc_id, TestCase.type, TestCase.category,
                TestCase.depth1, TestCase.depth2, TestCase.priority,
                TestCase.test_steps, TestCase.expected_result,
            )
        )
        .filter(TestResult.test_run_id == run.id)
        .all()
    )

    wb = Workbook()

    # ── Summary sheet ─────────────────────────────────────────────────────
    ws_summary = wb.active
    ws_summary.title = "Summary"

    dark_fill = PatternFill(start_color="2F3136", end_color="2F3136", fill_type="solid")
    header_font = Font(name="Malgun Gothic", bold=True, color="FFFFFF", size=10)
    title_font = Font(name="Malgun Gothic", bold=True, size=14)
    cell_font = Font(name="Malgun Gothic", size=10)
    thin_border = Border(
        left=Side(style="thin"), right=Side(style="thin"),
        top=Side(style="thin"), bottom=Side(style="thin"),
    )
    center = Alignment(horizontal="center", vertical="center")

    pass_fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
    fail_fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
    block_fill = PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid")

    ws_summary.merge_cells("B1:H1")
    ws_summary["B1"].value = f"{project.name} - Test Report"
    ws_summary["B1"].font = title_font

    ws_summary["B3"].value = f"Test Run: {run.name}"
    ws_summary["B4"].value = f"Version: {run.version or 'N/A'}  |  Environment: {run.environment or 'N/A'}  |  Round: {run.round}"

    # Summary table
    sum_headers = ["Total", "Pass", "Fail", "Block", "NA", "NS", "Pass Rate"]
    sum_values = [
        summary["total"], summary["passed"], summary["failed"],
        summary["blocked"], summary["na"], summary["ns"],
        f"{summary['pass_rate']}%",
    ]

    for i, (h, v) in enumerate(zip(sum_headers, sum_values)):
        col = i + 2
        hcell = ws_summary.cell(row=6, column=col, value=h)
        hcell.font = header_font
        hcell.fill = dark_fill
        hcell.alignment = center
        hcell.border = thin_border

        vcell = ws_summary.cell(row=7, column=col, value=v)
        vcell.font = cell_font
        vcell.alignment = center
        vcell.border = thin_border
        ws_summary.column_dimensions[get_column_letter(col)].width = 12

    # ── Results sheet ─────────────────────────────────────────────────────
    ws_results = wb.create_sheet("Results")

    res_headers = [
        "No", "TC ID", "Type", "Category", "Depth1", "Depth2",
        "Priority", "Steps", "Expected Result", "Result",
        "Actual Result", "Issue Link", "Remarks",
    ]
    res_widths = [6, 10, 10, 15, 18, 18, 10, 35, 35, 10, 35, 20, 20]

    for i, (h, w) in enumerate(zip(res_headers, res_widths)):
        col = i + 1
        cell = ws_results.cell(row=1, column=col, value=h)
        cell.font = header_font
        cell.fill = dark_fill
        cell.alignment = center
        cell.border = thin_border
        ws_results.column_dimensions[get_column_letter(col)].width = w

    for row_idx, r in enumerate(results, 2):
        tc = r.test_case
        result_val = r.result.value if hasattr(r.result, "value") else r.result
        row_values = [
            tc.no, tc.tc_id, tc.type, tc.category, tc.depth1, tc.depth2,
            tc.priority, tc.test_steps, tc.expected_result, result_val,
            r.actual_result, r.issue_link, r.remarks,
        ]
        for col_idx, val in enumerate(row_values, 1):
            cell = ws_results.cell(row=row_idx, column=col_idx, value=val)
            cell.font = cell_font
            cell.border = thin_border
            cell.alignment = Alignment(vertical="center", wrap_text=True)

            # Color the result column
            if col_idx == 10:
                cell.alignment = center
                if result_val == "PASS":
                    cell.fill = pass_fill
                elif result_val == "FAIL":
                    cell.fill = fail_fill
                elif result_val == "BLOCK":
                    cell.fill = block_fill

    ws_results.freeze_panes = "A2"

    output = io.BytesIO()
    wb.save(output)
    output.seek(0)

    from urllib.parse import quote
    filename = f"{project.name}_Report_R{run.round}.xlsx"
    encoded = quote(filename)
    return StreamingResponse(
        output,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename*=UTF-8''{encoded}"},
    )

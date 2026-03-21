"""TC Manager 보안 통합 테스트 (pytest 기반)

실행: cd backend && python -m pytest test_security.py -v
"""
import io
import os

import pytest
import requests

BASE = os.getenv("TEST_BASE_URL", "http://localhost:8008")


class TokenStore:
    admin: str = ""
    viewer: str = ""
    viewer_uid: int = 0


store = TokenStore()


def login(username, password):
    r = requests.post(f"{BASE}/api/auth/login", json={"username": username, "password": password})
    if r.status_code == 200:
        return r.json()["access_token"]
    return None


def auth(token):
    return {"Authorization": f"Bearer {token}"}


# ── Fixtures ──────────────────────────────────────────────────


@pytest.fixture(scope="session", autouse=True)
def setup_tokens():
    """세션 시작 시 토큰 준비"""
    store.admin = login("admin", os.getenv("TEST_ADMIN_PASSWORD", "test1234"))
    assert store.admin, "Admin login failed - is the server running?"

    # 테스트용 viewer 생성
    r = requests.post(f"{BASE}/api/auth/register", json={
        "username": "__sec_viewer__", "password": "viewer1234", "display_name": "Sec Viewer"
    })
    if r.status_code == 201:
        store.viewer_uid = r.json()["id"]
    else:
        r2 = requests.get(f"{BASE}/api/auth/users", headers=auth(store.admin))
        store.viewer_uid = next(
            (u["id"] for u in r2.json() if u["username"] == "__sec_viewer__"), 0
        )
    store.viewer = login("__sec_viewer__", "viewer1234")
    assert store.viewer, "Viewer login failed"


@pytest.fixture(scope="session")
def project_pair():
    """공개/비공개 프로젝트 쌍 생성, 테스트 종료 후 삭제"""
    h = auth(store.admin)
    r1 = requests.post(f"{BASE}/api/projects", headers=h, json={
        "name": "__sec_pub__", "is_private": False
    })
    r2 = requests.post(f"{BASE}/api/projects", headers=h, json={
        "name": "__sec_priv__", "is_private": True
    })
    pub_id = r1.json()["id"]
    priv_id = r2.json()["id"]
    yield pub_id, priv_id
    requests.delete(f"{BASE}/api/projects/{pub_id}", headers=h)
    requests.delete(f"{BASE}/api/projects/{priv_id}", headers=h)


# ── 1. 인증 ──────────────────────────────────────────────────


class TestAuth:
    def test_login_success(self):
        assert store.admin

    def test_wrong_password(self):
        r = requests.post(f"{BASE}/api/auth/login", json={"username": "admin", "password": "wrong"})
        assert r.status_code == 401

    def test_unknown_user(self):
        r = requests.post(f"{BASE}/api/auth/login", json={"username": "nouser", "password": "x"})
        assert r.status_code == 401

    def test_no_auth(self):
        r = requests.get(f"{BASE}/api/projects")
        assert r.status_code == 401

    def test_invalid_token(self):
        r = requests.get(f"{BASE}/api/projects", headers={"Authorization": "Bearer bad"})
        assert r.status_code == 401

    def test_me(self):
        r = requests.get(f"{BASE}/api/auth/me", headers=auth(store.admin))
        assert r.status_code == 200
        assert r.json()["username"] == "admin"


# ── 1-2. 비밀번호 변경 ─────────────────────────────────────────


class TestPasswordChange:
    def test_change_password_success(self):
        """정상 비밀번호 변경"""
        # viewer로 비밀번호 변경
        r = requests.put(f"{BASE}/api/auth/change-password", headers=auth(store.viewer), json={
            "current_password": "viewer1234",
            "new_password": "newpass1234",
        })
        assert r.status_code == 200
        # 변경된 비밀번호로 로그인
        token = login("__sec_viewer__", "newpass1234")
        assert token
        # 원래 비밀번호로 복원
        r2 = requests.put(f"{BASE}/api/auth/change-password", headers={"Authorization": f"Bearer {token}"}, json={
            "current_password": "newpass1234",
            "new_password": "viewer1234",
        })
        assert r2.status_code == 200
        store.viewer = login("__sec_viewer__", "viewer1234")

    def test_change_password_wrong_current(self):
        """현재 비밀번호 틀리면 400"""
        r = requests.put(f"{BASE}/api/auth/change-password", headers=auth(store.viewer), json={
            "current_password": "wrongpassword",
            "new_password": "newpass1234",
        })
        assert r.status_code == 400

    def test_change_password_same_as_current(self):
        """새 비밀번호가 현재와 동일하면 400"""
        r = requests.put(f"{BASE}/api/auth/change-password", headers=auth(store.viewer), json={
            "current_password": "viewer1234",
            "new_password": "viewer1234",
        })
        assert r.status_code == 400

    def test_change_password_too_short(self):
        """8자 미만 비밀번호 거부 (422)"""
        r = requests.put(f"{BASE}/api/auth/change-password", headers=auth(store.viewer), json={
            "current_password": "viewer1234",
            "new_password": "short",
        })
        assert r.status_code == 422

    def test_change_password_no_auth(self):
        """인증 없이 비밀번호 변경 시도 → 401"""
        r = requests.put(f"{BASE}/api/auth/change-password", json={
            "current_password": "viewer1234",
            "new_password": "newpass1234",
        })
        assert r.status_code == 401


# ── 1-3. 회원 관리 (admin 전용) ────────────────────────────────


class TestUserManagement:
    def test_list_users_as_admin(self):
        """admin은 사용자 목록 조회 가능"""
        r = requests.get(f"{BASE}/api/auth/users", headers=auth(store.admin))
        assert r.status_code == 200
        users = r.json()
        assert len(users) >= 2  # admin + viewer 최소

    def test_list_users_as_viewer_forbidden(self):
        """일반 유저는 사용자 목록 조회 불가"""
        r = requests.get(f"{BASE}/api/auth/users", headers=auth(store.viewer))
        assert r.status_code == 403

    def test_update_role_as_admin(self):
        """admin이 역할 변경"""
        r = requests.put(
            f"{BASE}/api/auth/users/{store.viewer_uid}/role",
            headers=auth(store.admin),
            json={"role": "qa_manager"},
        )
        assert r.status_code == 200
        assert r.json()["role"] == "qa_manager"
        # 원래 역할로 복원
        r2 = requests.put(
            f"{BASE}/api/auth/users/{store.viewer_uid}/role",
            headers=auth(store.admin),
            json={"role": "user"},
        )
        assert r2.status_code == 200

    def test_update_role_invalid(self):
        """존재하지 않는 역할로 변경 시 400"""
        r = requests.put(
            f"{BASE}/api/auth/users/{store.viewer_uid}/role",
            headers=auth(store.admin),
            json={"role": "superadmin"},
        )
        assert r.status_code == 400

    def test_update_role_as_viewer_forbidden(self):
        """일반 유저는 역할 변경 불가"""
        r = requests.put(
            f"{BASE}/api/auth/users/{store.viewer_uid}/role",
            headers=auth(store.viewer),
            json={"role": "admin"},
        )
        assert r.status_code == 403

    def test_update_role_nonexistent_user(self):
        """존재하지 않는 사용자 역할 변경 시 404"""
        r = requests.put(
            f"{BASE}/api/auth/users/999999/role",
            headers=auth(store.admin),
            json={"role": "user"},
        )
        assert r.status_code == 404

    def test_reset_password_as_admin(self):
        """admin이 비밀번호 초기화 → temp_password 반환"""
        r = requests.put(
            f"{BASE}/api/auth/users/{store.viewer_uid}/reset-password",
            headers=auth(store.admin),
        )
        assert r.status_code == 200
        assert "temp_password" in r.json()
        temp_pw = r.json()["temp_password"]
        assert len(temp_pw) == 12
        # 임시 비밀번호로 로그인 확인
        token = login("__sec_viewer__", temp_pw)
        assert token
        # viewer 비밀번호 복원
        requests.put(f"{BASE}/api/auth/change-password", headers={"Authorization": f"Bearer {token}"}, json={
            "current_password": temp_pw,
            "new_password": "viewer1234",
        })
        store.viewer = login("__sec_viewer__", "viewer1234")

    def test_reset_password_as_viewer_forbidden(self):
        """일반 유저는 비밀번호 초기화 불가"""
        r = requests.put(
            f"{BASE}/api/auth/users/{store.viewer_uid}/reset-password",
            headers=auth(store.viewer),
        )
        assert r.status_code == 403

    def test_check_username_exists(self):
        """존재하는 사용자명 → available: false"""
        r = requests.get(f"{BASE}/api/auth/check-username?username=admin")
        assert r.status_code == 200
        assert r.json()["available"] is False

    def test_check_username_available(self):
        """없는 사용자명 → available: true"""
        r = requests.get(f"{BASE}/api/auth/check-username?username=__nonexistent_user_12345__")
        assert r.status_code == 200
        assert r.json()["available"] is True

    def test_register_duplicate(self):
        """중복 사용자 등록 → 400"""
        r = requests.post(f"{BASE}/api/auth/register", json={
            "username": "admin",
            "password": "somepassword123",
            "display_name": "Dup Admin",
        })
        assert r.status_code == 400

    def test_register_short_password(self):
        """8자 미만 비밀번호로 등록 → 422"""
        r = requests.post(f"{BASE}/api/auth/register", json={
            "username": "__sec_short_pw__",
            "password": "short",
            "display_name": "Short PW",
        })
        assert r.status_code == 422


# ── 2. Rate Limiting ─────────────────────────────────────────


class TestRateLimiting:
    def test_rate_limit_on_failures(self):
        blocked = False
        for _ in range(15):
            r = requests.post(
                f"{BASE}/api/auth/login",
                json={"username": "__rl_test__", "password": "wrong"},
            )
            if r.status_code == 429:
                blocked = True
                break
        assert blocked, "Rate limiting did not trigger after 15 failures"

    def test_success_does_not_count(self):
        """성공 로그인은 실패 카운트에 누적되지 않아야 함"""
        # admin으로 연속 성공 로그인 - 차단되면 안 됨
        for _ in range(5):
            r = requests.post(
                f"{BASE}/api/auth/login",
                json={"username": "admin", "password": os.getenv("TEST_ADMIN_PASSWORD", "test1234")},
            )
            assert r.status_code == 200


# ── 3. 프로젝트 접근 권한 ────────────────────────────────────


class TestProjectAccess:
    def test_project_fields(self):
        r = requests.get(f"{BASE}/api/projects", headers=auth(store.admin))
        for p in r.json():
            assert "my_role" in p
            assert "is_private" in p

    def test_private_hidden_from_non_member(self, project_pair):
        _, priv_id = project_pair
        r = requests.get(f"{BASE}/api/projects", headers=auth(store.viewer))
        ids = [p["id"] for p in r.json()]
        assert priv_id not in ids

    def test_private_403_for_non_member(self, project_pair):
        _, priv_id = project_pair
        r = requests.get(f"{BASE}/api/projects/{priv_id}", headers=auth(store.viewer))
        assert r.status_code == 403

    def test_public_accessible(self, project_pair):
        pub_id, _ = project_pair
        r = requests.get(f"{BASE}/api/projects/{pub_id}", headers=auth(store.viewer))
        assert r.status_code == 200

    def test_viewer_cannot_create_tc(self, project_pair):
        pub_id, _ = project_pair
        r = requests.post(
            f"{BASE}/api/projects/{pub_id}/testcases",
            headers=auth(store.viewer),
            json={"no": 1, "tc_id": "TC-X"},
        )
        assert r.status_code == 403

    def test_private_accessible_after_member_add(self, project_pair):
        _, priv_id = project_pair
        h = auth(store.admin)
        requests.post(f"{BASE}/api/projects/{priv_id}/members", headers=h, json={
            "user_id": store.viewer_uid, "role": "tester"
        })
        r = requests.get(f"{BASE}/api/projects/{priv_id}", headers=auth(store.viewer))
        assert r.status_code == 200


# ── 4. 멤버 관리 ────────────────────────────────────────────


class TestMembers:
    def test_add_and_remove(self, project_pair):
        pub_id, _ = project_pair
        h = auth(store.admin)
        # 추가
        r = requests.post(f"{BASE}/api/projects/{pub_id}/members", headers=h, json={
            "user_id": store.viewer_uid, "role": "tester"
        })
        assert r.status_code == 201
        mid = r.json()["id"]
        # 중복 거부
        r2 = requests.post(f"{BASE}/api/projects/{pub_id}/members", headers=h, json={
            "user_id": store.viewer_uid, "role": "tester"
        })
        assert r2.status_code == 400
        # 역할 변경
        r3 = requests.put(f"{BASE}/api/projects/{pub_id}/members/{mid}", headers=h, json={"role": "admin"})
        assert r3.status_code == 200 and r3.json()["role"] == "admin"
        # 잘못된 역할
        r4 = requests.put(f"{BASE}/api/projects/{pub_id}/members/{mid}", headers=h, json={"role": "super"})
        assert r4.status_code == 400
        # 제거
        r5 = requests.delete(f"{BASE}/api/projects/{pub_id}/members/{mid}", headers=h)
        assert r5.status_code == 204

    def test_creator_cannot_be_removed(self, project_pair):
        pub_id, _ = project_pair
        h = auth(store.admin)
        r = requests.get(f"{BASE}/api/projects/{pub_id}/members", headers=h)
        creator = next((m for m in r.json() if m.get("username") == "admin"), None)
        if creator:
            r2 = requests.delete(f"{BASE}/api/projects/{pub_id}/members/{creator['id']}", headers=h)
            assert r2.status_code == 400


# ── 5. 파일 업로드 보안 ──────────────────────────────────────


class TestFileUpload:
    @pytest.fixture(autouse=True)
    def _setup_run(self, project_pair):
        """테스트용 TestRun & TestResult 준비"""
        pub_id, _ = project_pair
        h = auth(store.admin)
        # TC 생성
        r = requests.post(f"{BASE}/api/projects/{pub_id}/testcases", headers=h, json={
            "no": 888, "tc_id": "TC-SEC-001", "test_steps": "s"
        })
        self.tc_id = r.json().get("id") if r.status_code == 201 else None
        # TestRun 생성
        r = requests.post(f"{BASE}/api/projects/{pub_id}/testruns", headers=h, json={
            "name": "__sec_run__", "round": 1
        })
        self.run_id = r.json().get("id") if r.status_code == 201 else None
        if self.run_id:
            r2 = requests.get(f"{BASE}/api/projects/{pub_id}/testruns/{self.run_id}", headers=h)
            results = r2.json().get("results", [])
            self.tr_id = results[0]["id"] if results else None
        else:
            self.tr_id = None
        self.pub_id = pub_id
        self.h = h
        yield
        # 정리
        if self.run_id:
            requests.delete(f"{BASE}/api/projects/{pub_id}/testruns/{self.run_id}", headers=h)
        if self.tc_id:
            requests.delete(f"{BASE}/api/projects/{pub_id}/testcases/{self.tc_id}", headers=h)

    def test_exe_rejected(self):
        if not self.tr_id:
            pytest.skip("No test result")
        files = {"file": ("bad.exe", io.BytesIO(b"x"), "application/octet-stream")}
        r = requests.post(f"{BASE}/api/attachments/{self.tr_id}", headers=self.h, files=files)
        assert r.status_code == 400

    def test_no_extension_rejected(self):
        """확장자 없는 파일 업로드 차단"""
        if not self.tr_id:
            pytest.skip("No test result")
        files = {"file": ("noext", io.BytesIO(b"<script>alert(1)</script>"), "text/html")}
        r = requests.post(f"{BASE}/api/attachments/{self.tr_id}", headers=self.h, files=files)
        assert r.status_code == 400, f"Expected 400, got {r.status_code}"

    def test_png_allowed(self):
        if not self.tr_id:
            pytest.skip("No test result")
        files = {"file": ("ok.png", io.BytesIO(b"PNG"), "image/png")}
        r = requests.post(f"{BASE}/api/attachments/{self.tr_id}", headers=self.h, files=files)
        assert r.status_code == 200

    def test_attachment_project_access_check(self):
        """다른 프로젝트의 첨부파일에 접근 불가 (viewer가 비멤버 프로젝트)"""
        if not self.tr_id:
            pytest.skip("No test result")
        # admin으로 파일 업로드
        files = {"file": ("secret.pdf", io.BytesIO(b"secret"), "application/pdf")}
        r = requests.post(f"{BASE}/api/attachments/{self.tr_id}", headers=self.h, files=files)
        if r.status_code != 200:
            pytest.skip("Upload failed")
        att_id = r.json()["id"]
        # viewer가 다운로드 시도 — 공개 프로젝트이므로 접근 가능해야 함
        r2 = requests.get(f"{BASE}/api/attachments/download/{att_id}", headers=auth(store.viewer))
        assert r2.status_code == 200  # 공개 프로젝트라 viewer 접근 가능

    def test_xlsx_import_rejects_non_excel(self):
        files = {"file": ("test.txt", io.BytesIO(b"text"), "text/plain")}
        r = requests.post(
            f"{BASE}/api/projects/{self.pub_id}/testcases/import",
            headers=self.h, files=files,
        )
        assert r.status_code == 400


# ── 5-2. 첨부파일 프로젝트 경계 테스트 ──────────────────────


class TestAttachmentProjectBoundary:
    """비공개 프로젝트 첨부파일 접근 차단 및 크로스 프로젝트 경계 테스트"""

    @pytest.fixture(autouse=True)
    def _setup(self, project_pair):
        """비공개 프로젝트에 TC → TestRun → TestResult → Attachment 생성"""
        _, priv_id = project_pair
        h = auth(store.admin)
        # TC 생성
        r = requests.post(f"{BASE}/api/projects/{priv_id}/testcases", headers=h, json={
            "no": 999, "tc_id": "TC-PRIV-SEC", "test_steps": "s"
        })
        self.tc_id = r.json().get("id") if r.status_code == 201 else None
        # TestRun 생성
        r = requests.post(f"{BASE}/api/projects/{priv_id}/testruns", headers=h, json={
            "name": "__priv_sec_run__", "round": 1
        })
        self.run_id = r.json().get("id") if r.status_code == 201 else None
        self.tr_id = None
        self.att_id = None
        if self.run_id:
            r2 = requests.get(f"{BASE}/api/projects/{priv_id}/testruns/{self.run_id}", headers=h)
            results = r2.json().get("results", [])
            if results:
                self.tr_id = results[0]["id"]
                # 첨부파일 업로드
                files = {"file": ("priv.pdf", io.BytesIO(b"private-data"), "application/pdf")}
                r3 = requests.post(f"{BASE}/api/attachments/{self.tr_id}", headers=h, files=files)
                if r3.status_code == 200:
                    self.att_id = r3.json()["id"]
        self.priv_id = priv_id
        self.h = h
        yield
        # 정리
        if self.run_id:
            requests.delete(f"{BASE}/api/projects/{priv_id}/testruns/{self.run_id}", headers=h)
        if self.tc_id:
            requests.delete(f"{BASE}/api/projects/{priv_id}/testcases/{self.tc_id}", headers=h)

    def test_non_member_download_blocked(self):
        """비멤버가 비공개 프로젝트 첨부 다운로드 시 403"""
        if not self.att_id:
            pytest.skip("No attachment")
        # viewer를 비공개 프로젝트 멤버에서 제거 (있으면)
        h = auth(store.admin)
        members = requests.get(f"{BASE}/api/projects/{self.priv_id}/members", headers=h).json()
        for m in members:
            if m.get("user_id") == store.viewer_uid:
                requests.delete(f"{BASE}/api/projects/{self.priv_id}/members/{m['id']}", headers=h)
        # 비멤버 viewer가 다운로드 시도
        r = requests.get(f"{BASE}/api/attachments/download/{self.att_id}", headers=auth(store.viewer))
        assert r.status_code == 403, f"Expected 403, got {r.status_code}"

    def test_non_member_upload_blocked(self):
        """비멤버가 비공개 프로젝트에 첨부 업로드 시 403"""
        if not self.tr_id:
            pytest.skip("No test result")
        files = {"file": ("hack.png", io.BytesIO(b"PNG"), "image/png")}
        r = requests.post(f"{BASE}/api/attachments/{self.tr_id}", headers=auth(store.viewer), files=files)
        assert r.status_code == 403, f"Expected 403, got {r.status_code}"

    def test_non_member_delete_blocked(self):
        """비멤버가 비공개 프로젝트 첨부 삭제 시 403"""
        if not self.att_id:
            pytest.skip("No attachment")
        r = requests.delete(f"{BASE}/api/attachments/{self.att_id}", headers=auth(store.viewer))
        assert r.status_code == 403, f"Expected 403, got {r.status_code}"

    def test_non_member_list_blocked(self):
        """비멤버가 비공개 프로젝트 첨부 목록 조회 시 403"""
        if not self.tr_id:
            pytest.skip("No test result")
        r = requests.get(f"{BASE}/api/attachments/{self.tr_id}", headers=auth(store.viewer))
        assert r.status_code == 403, f"Expected 403, got {r.status_code}"


# ── 6. 코드 보안 검증 ───────────────────────────────────────


class TestCodeSecurity:
    def _read(self, relpath):
        base = os.path.dirname(os.path.abspath(__file__))
        with open(os.path.join(base, relpath), "r", encoding="utf-8") as f:
            return f.read()

    def test_no_hardcoded_secret(self):
        assert "dev-secret-key" not in self._read("auth.py")

    def test_random_key_generation(self):
        assert "secrets.token_urlsafe" in self._read("auth.py")

    def test_cors_no_wildcard_methods(self):
        assert 'allow_methods=["*"]' not in self._read("main.py")

    def test_cors_no_wildcard_headers(self):
        assert 'allow_headers=["*"]' not in self._read("main.py")

    def test_path_traversal_defense(self):
        assert "_safe_filepath" in self._read("routes/attachments.py")

    def test_file_size_limit(self):
        assert "MAX_FILE_SIZE" in self._read("routes/attachments.py")

    def test_extension_whitelist(self):
        assert "ALLOWED_EXTENSIONS" in self._read("routes/attachments.py")

    def test_no_extension_rejected_in_code(self):
        code = self._read("routes/attachments.py")
        assert "File must have an extension" in code

    def test_safe_content_type(self):
        assert "SAFE_CONTENT_TYPES" in self._read("routes/attachments.py")

    def test_rate_limit_failure_only(self):
        code = self._read("routes/auth.py")
        assert "_record_failure" in code
        assert "_clear_failures" in code

    def test_excel_import_size_limit(self):
        assert "MAX_IMPORT_SIZE" in self._read("routes/testcases.py")

    def test_dompurify_installed(self):
        fe = os.path.join(os.path.dirname(os.path.abspath(__file__)), "..", "frontend")
        assert os.path.exists(os.path.join(fe, "node_modules", "dompurify"))

    def test_dompurify_applied(self):
        fe = os.path.join(os.path.dirname(os.path.abspath(__file__)), "..", "frontend")
        mc = os.path.join(fe, "src", "components", "MarkdownCell.tsx")
        with open(mc, "r", encoding="utf-8") as f:
            assert "DOMPurify.sanitize" in f.read()

    def test_attachment_project_check_in_code(self):
        code = self._read("routes/attachments.py")
        assert "_check_attachment_access" in code
        assert "_get_project_id_from_test_result" in code

    def test_rate_limit_memory_protection(self):
        code = self._read("routes/auth.py")
        assert "_MAX_RATE_LIMIT_KEYS" in code
        assert "_purge_expired_keys" in code

    def test_pinned_dependencies(self):
        assert "fastapi==" in self._read("requirements.txt")

    def test_dashboard_uses_check_project_access(self):
        code = self._read("routes/dashboard.py")
        assert "check_project_access" in code
        assert 'Depends(get_current_user)' not in code

    def test_reports_uses_check_project_access(self):
        code = self._read("routes/reports.py")
        assert "check_project_access" in code
        assert 'current_user: User = Depends(get_current_user)' not in code

    def test_search_filters_by_project_access(self):
        code = self._read("routes/search.py")
        assert "_accessible_project_ids" in code
        assert "project_id.in_" in code
        assert "UserRole.admin" in code
        assert "created_by" in code

    def test_overview_filters_private_projects(self):
        code = self._read("routes/overview.py")
        assert "is_private" in code
        assert "ProjectMember" in code
        assert "UserRole.admin" in code
        assert "created_by" in code


class TestBrokenAccessControl:
    """비공개 프로젝트에 대한 접근 통제 통합 테스트."""

    def test_dashboard_blocked_for_non_member(self, project_pair):
        _, priv_id = project_pair
        h = auth(store.viewer)
        r = requests.get(f"{BASE}/api/projects/{priv_id}/dashboard/summary", headers=h)
        assert r.status_code == 403

    def test_reports_blocked_for_non_member(self, project_pair):
        _, priv_id = project_pair
        h = auth(store.viewer)
        r = requests.get(f"{BASE}/api/projects/{priv_id}/reports?run_id=1", headers=h)
        assert r.status_code == 403

    def test_search_excludes_private_project(self, project_pair):
        pub_id, priv_id = project_pair
        h_admin = auth(store.admin)
        # 비공개 프로젝트에 TC 생성
        cr = requests.post(
            f"{BASE}/api/projects/{priv_id}/testcases",
            headers=h_admin,
            json={"no": 9001, "tc_id": "__sec_search_priv__", "category": "SecTest"},
        )
        assert cr.status_code in (200, 201), f"TC creation failed: {cr.status_code} {cr.text}"
        # viewer로 검색 시 비공개 TC가 나오지 않아야 함
        h = auth(store.viewer)
        r = requests.get(f"{BASE}/api/search?q=__sec_search_priv__", headers=h)
        assert r.status_code == 200
        ids = [tc["tc_id"] for tc in r.json()]
        assert "__sec_search_priv__" not in ids

    def test_overview_excludes_private_project(self, project_pair):
        _, priv_id = project_pair
        h = auth(store.viewer)
        r = requests.get(f"{BASE}/api/dashboard/overview", headers=h)
        assert r.status_code == 200
        project_ids = [p["id"] for p in r.json()["projects"]]
        assert priv_id not in project_ids

    # ── admin/creator 접근 보장 테스트 ──

    def test_admin_can_access_private_dashboard(self, project_pair):
        """시스템 admin은 비공개 프로젝트 dashboard에 접근 가능해야 함."""
        _, priv_id = project_pair
        h = auth(store.admin)
        r = requests.get(f"{BASE}/api/projects/{priv_id}/dashboard/summary", headers=h)
        assert r.status_code == 200

    def test_admin_can_access_private_reports(self, project_pair):
        """시스템 admin은 비공개 프로젝트 reports에 접근 가능해야 함."""
        _, priv_id = project_pair
        h = auth(store.admin)
        # run_id=999999 → 400 (run not found), not 403
        r = requests.get(f"{BASE}/api/projects/{priv_id}/reports?run_id=999999", headers=h)
        assert r.status_code != 403

    def test_admin_search_includes_private_project(self, project_pair):
        """시스템 admin은 비공개 프로젝트 TC도 검색 가능해야 함."""
        _, priv_id = project_pair
        h = auth(store.admin)
        # TC 직접 생성
        cr = requests.post(
            f"{BASE}/api/projects/{priv_id}/testcases",
            headers=h,
            json={"no": 9002, "tc_id": "__sec_admin_search__", "category": "SecAdminTest"},
        )
        assert cr.status_code in (200, 201), f"TC creation failed: {cr.status_code} {cr.text}"
        r = requests.get(f"{BASE}/api/search?q=__sec_admin_search__", headers=h)
        assert r.status_code == 200
        ids = [tc["tc_id"] for tc in r.json()]
        assert "__sec_admin_search__" in ids

    def test_admin_overview_includes_private_project(self, project_pair):
        """시스템 admin은 overview에서 비공개 프로젝트도 볼 수 있어야 함."""
        _, priv_id = project_pair
        h = auth(store.admin)
        r = requests.get(f"{BASE}/api/dashboard/overview", headers=h)
        assert r.status_code == 200
        project_ids = [p["id"] for p in r.json()["projects"]]
        assert priv_id in project_ids


# ── 7. 시트 관리 ──────────────────────────────────────────────


class TestSheetManagement:
    """시트 생성/목록/삭제 관련 테스트 (12개)."""

    def test_create_sheet_success(self, project_pair):
        """시트 생성 성공"""
        pub_id, _ = project_pair
        h = auth(store.admin)
        r = requests.post(f"{BASE}/api/projects/{pub_id}/testcases/sheets", headers=h, json={"name": "__sec_sheet_1__"})
        assert r.status_code == 200
        assert r.json()["name"] == "__sec_sheet_1__"
        assert r.json()["tc_count"] == 0
        # cleanup
        requests.delete(f"{BASE}/api/projects/{pub_id}/testcases/sheets/__sec_sheet_1__", headers=h)

    def test_create_sheet_duplicate(self, project_pair):
        """중복 시트 이름 → 400"""
        pub_id, _ = project_pair
        h = auth(store.admin)
        requests.post(f"{BASE}/api/projects/{pub_id}/testcases/sheets", headers=h, json={"name": "__sec_sheet_dup__"})
        r = requests.post(f"{BASE}/api/projects/{pub_id}/testcases/sheets", headers=h, json={"name": "__sec_sheet_dup__"})
        assert r.status_code == 400
        # cleanup
        requests.delete(f"{BASE}/api/projects/{pub_id}/testcases/sheets/__sec_sheet_dup__", headers=h)

    def test_create_sheet_empty_name(self, project_pair):
        """빈 시트 이름 → 400"""
        pub_id, _ = project_pair
        h = auth(store.admin)
        r = requests.post(f"{BASE}/api/projects/{pub_id}/testcases/sheets", headers=h, json={"name": "   "})
        assert r.status_code == 400

    def test_create_sheet_viewer_forbidden(self, project_pair):
        """viewer는 시트 생성 불가 → 403"""
        pub_id, _ = project_pair
        r = requests.post(f"{BASE}/api/projects/{pub_id}/testcases/sheets", headers=auth(store.viewer), json={"name": "__sec_sheet_v__"})
        assert r.status_code == 403

    def test_list_sheets_with_count(self, project_pair):
        """시트 목록에서 TC 카운트 정확성"""
        pub_id, _ = project_pair
        h = auth(store.admin)
        # 시트 생성 + TC 추가
        requests.post(f"{BASE}/api/projects/{pub_id}/testcases/sheets", headers=h, json={"name": "__sec_sheet_cnt__"})
        tc_r = requests.post(f"{BASE}/api/projects/{pub_id}/testcases", headers=h, json={
            "no": 7001, "tc_id": "TC-SHEET-CNT-001", "test_steps": "s", "sheet_name": "__sec_sheet_cnt__"
        })
        tc_id = tc_r.json().get("id") if tc_r.status_code == 201 else None

        r = requests.get(f"{BASE}/api/projects/{pub_id}/testcases/sheets", headers=h)
        assert r.status_code == 200
        sheet = next((s for s in r.json() if s["name"] == "__sec_sheet_cnt__"), None)
        assert sheet is not None
        assert sheet["tc_count"] >= 1

        # cleanup
        if tc_id:
            requests.delete(f"{BASE}/api/projects/{pub_id}/testcases/{tc_id}", headers=h)
        requests.delete(f"{BASE}/api/projects/{pub_id}/testcases/sheets/__sec_sheet_cnt__", headers=h)

    def test_list_sheets_empty_project(self, project_pair):
        """시트가 없는 프로젝트에서 빈 배열 반환"""
        _, priv_id = project_pair
        h = auth(store.admin)
        r = requests.get(f"{BASE}/api/projects/{priv_id}/testcases/sheets", headers=h)
        assert r.status_code == 200
        assert isinstance(r.json(), list)

    def test_list_sheets_sort_order(self, project_pair):
        """시트가 생성 순서대로 정렬"""
        pub_id, _ = project_pair
        h = auth(store.admin)
        requests.post(f"{BASE}/api/projects/{pub_id}/testcases/sheets", headers=h, json={"name": "__sec_sort_a__"})
        requests.post(f"{BASE}/api/projects/{pub_id}/testcases/sheets", headers=h, json={"name": "__sec_sort_b__"})

        r = requests.get(f"{BASE}/api/projects/{pub_id}/testcases/sheets", headers=h)
        names = [s["name"] for s in r.json()]
        if "__sec_sort_a__" in names and "__sec_sort_b__" in names:
            assert names.index("__sec_sort_a__") < names.index("__sec_sort_b__")

        # cleanup
        requests.delete(f"{BASE}/api/projects/{pub_id}/testcases/sheets/__sec_sort_a__", headers=h)
        requests.delete(f"{BASE}/api/projects/{pub_id}/testcases/sheets/__sec_sort_b__", headers=h)

    def test_delete_sheet_soft_deletes_tc(self, project_pair):
        """시트 삭제 시 소속 TC가 소프트 삭제됨"""
        pub_id, _ = project_pair
        h = auth(store.admin)
        requests.post(f"{BASE}/api/projects/{pub_id}/testcases/sheets", headers=h, json={"name": "__sec_del_sheet__"})
        tc_r = requests.post(f"{BASE}/api/projects/{pub_id}/testcases", headers=h, json={
            "no": 7010, "tc_id": "TC-DEL-SHEET-001", "test_steps": "s", "sheet_name": "__sec_del_sheet__"
        })
        tc_id = tc_r.json().get("id") if tc_r.status_code == 201 else None

        r = requests.delete(f"{BASE}/api/projects/{pub_id}/testcases/sheets/__sec_del_sheet__", headers=h)
        assert r.status_code == 200
        assert r.json()["deleted"] >= 1

        # TC가 목록에서 보이지 않아야 함
        tcs = requests.get(f"{BASE}/api/projects/{pub_id}/testcases", headers=h).json()
        tc_ids = [tc["tc_id"] for tc in tcs]
        assert "TC-DEL-SHEET-001" not in tc_ids

        # cleanup: hard delete via restore + delete
        if tc_id:
            requests.post(f"{BASE}/api/projects/{pub_id}/testcases/{tc_id}/restore", headers=h)
            requests.delete(f"{BASE}/api/projects/{pub_id}/testcases/{tc_id}", headers=h)

    def test_delete_sheet_returns_count(self, project_pair):
        """시트 삭제 응답에 삭제된 TC 수 포함"""
        pub_id, _ = project_pair
        h = auth(store.admin)
        requests.post(f"{BASE}/api/projects/{pub_id}/testcases/sheets", headers=h, json={"name": "__sec_del_cnt__"})
        # TC 없이 삭제
        r = requests.delete(f"{BASE}/api/projects/{pub_id}/testcases/sheets/__sec_del_cnt__", headers=h)
        assert r.status_code == 200
        assert r.json()["deleted"] == 0

    def test_delete_sheet_viewer_forbidden(self, project_pair):
        """viewer는 시트 삭제 불가 → 403"""
        pub_id, _ = project_pair
        h = auth(store.admin)
        requests.post(f"{BASE}/api/projects/{pub_id}/testcases/sheets", headers=h, json={"name": "__sec_del_vw__"})
        r = requests.delete(f"{BASE}/api/projects/{pub_id}/testcases/sheets/__sec_del_vw__", headers=auth(store.viewer))
        assert r.status_code == 403
        # cleanup
        requests.delete(f"{BASE}/api/projects/{pub_id}/testcases/sheets/__sec_del_vw__", headers=h)

    def test_delete_nonexistent_sheet(self, project_pair):
        """존재하지 않는 시트 삭제 → deleted: 0"""
        pub_id, _ = project_pair
        h = auth(store.admin)
        r = requests.delete(f"{BASE}/api/projects/{pub_id}/testcases/sheets/__nonexistent_sheet__", headers=h)
        assert r.status_code == 200
        assert r.json()["deleted"] == 0

    def test_delete_sheet_name_in_response(self, project_pair):
        """시트 삭제 응답에 시트 이름 포함"""
        pub_id, _ = project_pair
        h = auth(store.admin)
        requests.post(f"{BASE}/api/projects/{pub_id}/testcases/sheets", headers=h, json={"name": "__sec_del_name__"})
        r = requests.delete(f"{BASE}/api/projects/{pub_id}/testcases/sheets/__sec_del_name__", headers=h)
        assert r.status_code == 200
        assert r.json()["sheet"] == "__sec_del_name__"


# ── 8. 일괄 삭제 ─────────────────────────────────────────────


class TestBulkDelete:
    """TC 일괄 삭제 관련 테스트 (5개)."""

    def test_bulk_delete_success(self, project_pair):
        """여러 TC 일괄 소프트 삭제 성공"""
        pub_id, _ = project_pair
        h = auth(store.admin)
        ids = []
        for i in range(3):
            r = requests.post(f"{BASE}/api/projects/{pub_id}/testcases", headers=h, json={
                "no": 8001 + i, "tc_id": f"TC-BULK-{i}", "test_steps": "s"
            })
            if r.status_code == 201:
                ids.append(r.json()["id"])

        id_str = ",".join(str(x) for x in ids)
        r = requests.delete(f"{BASE}/api/projects/{pub_id}/testcases/bulk?ids={id_str}", headers=h)
        assert r.status_code == 200
        assert r.json()["deleted"] == len(ids)

        # cleanup
        for tc_id in ids:
            requests.post(f"{BASE}/api/projects/{pub_id}/testcases/{tc_id}/restore", headers=h)
            requests.delete(f"{BASE}/api/projects/{pub_id}/testcases/{tc_id}", headers=h)

    def test_bulk_delete_viewer_forbidden(self, project_pair):
        """viewer는 일괄 삭제 불가 → 403"""
        pub_id, _ = project_pair
        r = requests.delete(f"{BASE}/api/projects/{pub_id}/testcases/bulk?ids=1,2,3", headers=auth(store.viewer))
        assert r.status_code == 403

    def test_bulk_delete_cross_project_defense(self, project_pair):
        """다른 프로젝트의 TC ID로 삭제 시도 → deleted: 0"""
        pub_id, priv_id = project_pair
        h = auth(store.admin)
        # pub 프로젝트에 TC 생성
        r = requests.post(f"{BASE}/api/projects/{pub_id}/testcases", headers=h, json={
            "no": 8010, "tc_id": "TC-CROSS-BULK", "test_steps": "s"
        })
        tc_id = r.json()["id"] if r.status_code == 201 else None
        assert tc_id

        # priv 프로젝트에서 해당 TC 삭제 시도
        r2 = requests.delete(f"{BASE}/api/projects/{priv_id}/testcases/bulk?ids={tc_id}", headers=h)
        assert r2.status_code == 200
        assert r2.json()["deleted"] == 0

        # cleanup
        requests.delete(f"{BASE}/api/projects/{pub_id}/testcases/{tc_id}", headers=h)

    def test_bulk_delete_already_deleted(self, project_pair):
        """이미 삭제된 TC를 다시 삭제 시도 → deleted: 0"""
        pub_id, _ = project_pair
        h = auth(store.admin)
        r = requests.post(f"{BASE}/api/projects/{pub_id}/testcases", headers=h, json={
            "no": 8020, "tc_id": "TC-ALREADY-DEL", "test_steps": "s"
        })
        tc_id = r.json()["id"] if r.status_code == 201 else None
        assert tc_id

        # 첫 번째 삭제
        requests.delete(f"{BASE}/api/projects/{pub_id}/testcases/{tc_id}", headers=h)
        # 두 번째 삭제 (이미 삭제됨)
        r2 = requests.delete(f"{BASE}/api/projects/{pub_id}/testcases/bulk?ids={tc_id}", headers=h)
        assert r2.status_code == 200
        assert r2.json()["deleted"] == 0

        # cleanup
        requests.post(f"{BASE}/api/projects/{pub_id}/testcases/{tc_id}/restore", headers=h)
        requests.delete(f"{BASE}/api/projects/{pub_id}/testcases/{tc_id}", headers=h)

    def test_bulk_delete_empty_ids(self, project_pair):
        """빈 ID 목록으로 삭제 시도"""
        pub_id, _ = project_pair
        h = auth(store.admin)
        r = requests.delete(f"{BASE}/api/projects/{pub_id}/testcases/bulk?ids=", headers=h)
        # 빈 문자열 → deleted: 0 또는 422
        assert r.status_code in (200, 422)


# ── 9. 임포트 Upsert ─────────────────────────────────────────


class TestImportUpsert:
    """엑셀 임포트 관련 테스트 (5개)."""

    @staticmethod
    def _make_excel(rows, sheet_name="TestSheet"):
        """openpyxl로 간단한 엑셀 파일을 메모리에서 생성한다.

        rows: list of dict with keys like tc_id, no, test_steps, etc.
        """
        from openpyxl import Workbook as _Wb
        wb = _Wb()
        ws = wb.active
        ws.title = sheet_name
        headers = ["No", "TC ID", "Category", "Test Steps", "Expected Result"]
        for col, h in enumerate(headers, 1):
            ws.cell(row=1, column=col, value=h)
        for row_idx, row_data in enumerate(rows, 2):
            ws.cell(row=row_idx, column=1, value=row_data.get("no", row_idx - 1))
            ws.cell(row=row_idx, column=2, value=row_data.get("tc_id", f"TC-IMP-{row_idx}"))
            ws.cell(row=row_idx, column=3, value=row_data.get("category", "ImportTest"))
            ws.cell(row=row_idx, column=4, value=row_data.get("test_steps", "step"))
            ws.cell(row=row_idx, column=5, value=row_data.get("expected_result", "ok"))
        buf = io.BytesIO()
        wb.save(buf)
        buf.seek(0)
        return buf

    def test_import_upsert_updated_count(self, project_pair):
        """같은 tc_id로 2번 임포트 시 updated 카운트 확인"""
        pub_id, _ = project_pair
        h = auth(store.admin)
        rows = [{"no": 9001, "tc_id": "TC-UPSERT-001", "test_steps": "v1"}]

        # 1차 임포트
        buf1 = self._make_excel(rows)
        files1 = {"file": ("test.xlsx", buf1, "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")}
        r1 = requests.post(f"{BASE}/api/projects/{pub_id}/testcases/import", headers=h, files=files1)
        assert r1.status_code == 201
        assert r1.json()["created"] >= 1

        # 2차 임포트 (같은 tc_id)
        rows2 = [{"no": 9001, "tc_id": "TC-UPSERT-001", "test_steps": "v2"}]
        buf2 = self._make_excel(rows2)
        files2 = {"file": ("test.xlsx", buf2, "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")}
        r2 = requests.post(f"{BASE}/api/projects/{pub_id}/testcases/import", headers=h, files=files2)
        assert r2.status_code == 201
        assert r2.json()["updated"] >= 1

        # cleanup: TC 삭제
        tcs = requests.get(f"{BASE}/api/projects/{pub_id}/testcases?search=TC-UPSERT-001", headers=h).json()
        for tc in tcs:
            requests.delete(f"{BASE}/api/projects/{pub_id}/testcases/{tc['id']}", headers=h)
        # 시트 레코드 삭제
        requests.delete(f"{BASE}/api/projects/{pub_id}/testcases/sheets/TestSheet", headers=h)

    def test_preview_existing_count(self, project_pair):
        """preview API에서 existing 수 정확성 확인"""
        pub_id, _ = project_pair
        h = auth(store.admin)
        rows = [{"no": 9010, "tc_id": "TC-PREVIEW-001", "test_steps": "s"}]
        buf = self._make_excel(rows, sheet_name="PreviewSheet")
        files = {"file": ("test.xlsx", buf, "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")}

        # 먼저 임포트
        r_imp = requests.post(
            f"{BASE}/api/projects/{pub_id}/testcases/import?sheet_names=PreviewSheet",
            headers=h, files=files,
        )
        assert r_imp.status_code == 201

        # preview
        buf2 = self._make_excel(rows, sheet_name="PreviewSheet")
        files2 = {"file": ("test.xlsx", buf2, "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")}
        r = requests.post(f"{BASE}/api/projects/{pub_id}/testcases/import/preview", headers=h, files=files2)
        assert r.status_code == 200
        sheet_info = next((s for s in r.json()["sheets"] if s["name"] == "PreviewSheet"), None)
        assert sheet_info is not None
        assert sheet_info["existing"] >= 1

        # cleanup
        tcs = requests.get(f"{BASE}/api/projects/{pub_id}/testcases?search=TC-PREVIEW-001", headers=h).json()
        for tc in tcs:
            requests.delete(f"{BASE}/api/projects/{pub_id}/testcases/{tc['id']}", headers=h)
        requests.delete(f"{BASE}/api/projects/{pub_id}/testcases/sheets/PreviewSheet", headers=h)

    def test_import_creates_sheet_record(self, project_pair):
        """임포트 시 TestCaseSheet 레코드 자동 생성 확인"""
        pub_id, _ = project_pair
        h = auth(store.admin)
        rows = [{"no": 9020, "tc_id": "TC-SHEETREC-001", "test_steps": "s"}]
        buf = self._make_excel(rows, sheet_name="AutoCreatedSheet")
        files = {"file": ("test.xlsx", buf, "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")}

        r = requests.post(
            f"{BASE}/api/projects/{pub_id}/testcases/import?sheet_names=AutoCreatedSheet",
            headers=h, files=files,
        )
        assert r.status_code == 201

        # 시트 목록에 나타나야 함
        sheets = requests.get(f"{BASE}/api/projects/{pub_id}/testcases/sheets", headers=h).json()
        sheet_names = [s["name"] for s in sheets]
        assert "AutoCreatedSheet" in sheet_names

        # cleanup
        tcs = requests.get(f"{BASE}/api/projects/{pub_id}/testcases?sheet_name=AutoCreatedSheet", headers=h).json()
        for tc in tcs:
            requests.delete(f"{BASE}/api/projects/{pub_id}/testcases/{tc['id']}", headers=h)
        requests.delete(f"{BASE}/api/projects/{pub_id}/testcases/sheets/AutoCreatedSheet", headers=h)

    def test_import_specific_sheet_only(self, project_pair):
        """sheet_names 파라미터로 특정 시트만 선택 임포트"""
        pub_id, _ = project_pair
        h = auth(store.admin)

        # 2개 시트 엑셀 생성
        from openpyxl import Workbook as _Wb
        wb = _Wb()
        ws1 = wb.active
        ws1.title = "SheetA"
        for col, hdr in enumerate(["No", "TC ID", "Test Steps"], 1):
            ws1.cell(row=1, column=col, value=hdr)
        ws1.cell(row=2, column=1, value=1)
        ws1.cell(row=2, column=2, value="TC-SEL-A")
        ws1.cell(row=2, column=3, value="step a")

        ws2 = wb.create_sheet("SheetB")
        for col, hdr in enumerate(["No", "TC ID", "Test Steps"], 1):
            ws2.cell(row=1, column=col, value=hdr)
        ws2.cell(row=2, column=1, value=1)
        ws2.cell(row=2, column=2, value="TC-SEL-B")
        ws2.cell(row=2, column=3, value="step b")

        buf = io.BytesIO()
        wb.save(buf)
        buf.seek(0)

        files = {"file": ("multi.xlsx", buf, "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")}
        r = requests.post(
            f"{BASE}/api/projects/{pub_id}/testcases/import?sheet_names=SheetA",
            headers=h, files=files,
        )
        assert r.status_code == 201
        # SheetA만 임포트됨
        sheet_results = r.json().get("sheets", [])
        imported_sheets = [s["sheet"] for s in sheet_results]
        assert "SheetA" in imported_sheets
        assert "SheetB" not in imported_sheets

        # cleanup
        tcs = requests.get(f"{BASE}/api/projects/{pub_id}/testcases?search=TC-SEL-A", headers=h).json()
        for tc in tcs:
            requests.delete(f"{BASE}/api/projects/{pub_id}/testcases/{tc['id']}", headers=h)
        requests.delete(f"{BASE}/api/projects/{pub_id}/testcases/sheets/SheetA", headers=h)

    def test_import_viewer_forbidden(self, project_pair):
        """viewer는 임포트 불가 → 403"""
        pub_id, _ = project_pair
        rows = [{"no": 9099, "tc_id": "TC-IMP-VIEWER", "test_steps": "s"}]
        buf = self._make_excel(rows)
        files = {"file": ("test.xlsx", buf, "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")}
        r = requests.post(f"{BASE}/api/projects/{pub_id}/testcases/import", headers=auth(store.viewer), files=files)
        assert r.status_code == 403


# ── 10. deleted_at 필터 ───────────────────────────────────────


class TestDeletedAtFilter:
    """소프트 삭제된 TC의 필터링 관련 테스트 (5개)."""

    @pytest.fixture(autouse=True)
    def _setup_tc(self, project_pair):
        """테스트용 TC 생성 → 소프트 삭제"""
        pub_id, _ = project_pair
        h = auth(store.admin)
        r = requests.post(f"{BASE}/api/projects/{pub_id}/testcases", headers=h, json={
            "no": 9900, "tc_id": "TC-SOFTDEL-001", "test_steps": "s", "category": "SoftDelTest"
        })
        self.tc_id = r.json().get("id") if r.status_code == 201 else None
        self.pub_id = pub_id
        self.h = h
        yield
        # cleanup: restore + hard-check
        if self.tc_id:
            requests.post(f"{BASE}/api/projects/{pub_id}/testcases/{self.tc_id}/restore", headers=h)
            requests.delete(f"{BASE}/api/projects/{pub_id}/testcases/{self.tc_id}", headers=h)

    def test_deleted_tc_excluded_from_list(self):
        """소프트 삭제된 TC가 GET /testcases에서 제외"""
        if not self.tc_id:
            pytest.skip("TC not created")
        # 소프트 삭제
        requests.delete(f"{BASE}/api/projects/{self.pub_id}/testcases/{self.tc_id}", headers=self.h)
        tcs = requests.get(f"{BASE}/api/projects/{self.pub_id}/testcases", headers=self.h).json()
        ids = [tc["id"] for tc in tcs]
        assert self.tc_id not in ids

    def test_deleted_tc_excluded_from_dashboard_summary(self):
        """소프트 삭제된 TC가 dashboard summary total에서 제외"""
        if not self.tc_id:
            pytest.skip("TC not created")
        # 삭제 전 total
        r_before = requests.get(f"{BASE}/api/projects/{self.pub_id}/dashboard/summary", headers=self.h)
        total_before = r_before.json()["total"]
        # 소프트 삭제
        requests.delete(f"{BASE}/api/projects/{self.pub_id}/testcases/{self.tc_id}", headers=self.h)
        r_after = requests.get(f"{BASE}/api/projects/{self.pub_id}/dashboard/summary", headers=self.h)
        total_after = r_after.json()["total"]
        assert total_after < total_before

    def test_deleted_tc_excluded_from_overview(self):
        """소프트 삭제된 TC가 overview total_tc에서 제외"""
        if not self.tc_id:
            pytest.skip("TC not created")
        r_before = requests.get(f"{BASE}/api/dashboard/overview", headers=self.h)
        total_before = r_before.json()["summary"]["total_tc"]
        # 소프트 삭제
        requests.delete(f"{BASE}/api/projects/{self.pub_id}/testcases/{self.tc_id}", headers=self.h)
        r_after = requests.get(f"{BASE}/api/dashboard/overview", headers=self.h)
        total_after = r_after.json()["summary"]["total_tc"]
        assert total_after < total_before

    def test_restore_brings_back_tc(self):
        """restore 후 TC가 다시 목록에 나타남"""
        if not self.tc_id:
            pytest.skip("TC not created")
        # 소프트 삭제
        requests.delete(f"{BASE}/api/projects/{self.pub_id}/testcases/{self.tc_id}", headers=self.h)
        # restore
        r = requests.post(f"{BASE}/api/projects/{self.pub_id}/testcases/{self.tc_id}/restore", headers=self.h)
        assert r.status_code == 200
        # 목록에 다시 나타남
        tcs = requests.get(f"{BASE}/api/projects/{self.pub_id}/testcases", headers=self.h).json()
        ids = [tc["id"] for tc in tcs]
        assert self.tc_id in ids

    def test_testrun_excludes_deleted_tc(self):
        """TestRun 생성 시 삭제된 TC가 TestResult에 포함되지 않음"""
        if not self.tc_id:
            pytest.skip("TC not created")
        # 소프트 삭제
        requests.delete(f"{BASE}/api/projects/{self.pub_id}/testcases/{self.tc_id}", headers=self.h)
        # TestRun 생성
        r = requests.post(f"{BASE}/api/projects/{self.pub_id}/testruns", headers=self.h, json={
            "name": "__sec_del_run__", "round": 1
        })
        if r.status_code != 201:
            pytest.skip("TestRun creation failed")
        run_id = r.json()["id"]
        # TestResult에서 삭제된 TC의 결과가 없어야 함
        run_detail = requests.get(f"{BASE}/api/projects/{self.pub_id}/testruns/{run_id}", headers=self.h).json()
        result_tc_ids = [res["test_case_id"] for res in run_detail.get("results", [])]
        assert self.tc_id not in result_tc_ids
        # cleanup
        requests.delete(f"{BASE}/api/projects/{self.pub_id}/testruns/{run_id}", headers=self.h)


# ── 11. 코드 보안 검증 (신규) ─────────────────────────────────


class TestCodeSecurityNew:
    """deleted_at 필터 코드 존재 확인 테스트 (3개)."""

    def _read(self, relpath):
        base = os.path.dirname(os.path.abspath(__file__))
        with open(os.path.join(base, relpath), "r", encoding="utf-8") as f:
            return f.read()

    def test_overview_has_deleted_at_filter(self):
        """overview.py에 deleted_at.is_(None) 필터 존재 확인"""
        code = self._read("routes/overview.py")
        assert "deleted_at.is_(None)" in code

    def test_dashboard_has_deleted_at_filter(self):
        """dashboard.py에 deleted_at.is_(None) 필터 존재 확인"""
        code = self._read("routes/dashboard.py")
        assert "deleted_at.is_(None)" in code

    def test_testruns_has_deleted_at_filter(self):
        """testruns.py에 deleted_at.is_(None) 필터 존재 확인"""
        code = self._read("routes/testruns.py")
        assert "deleted_at.is_(None)" in code


# ── 12. 검색 API ──────────────────────────────────────────────


class TestSearch:
    """GET /api/search 엔드포인트 테스트 (6개)."""

    def test_search_returns_results(self, project_pair):
        """공개 프로젝트에 TC 생성 후 검색하면 결과에 포함"""
        pub_id, _ = project_pair
        h = auth(store.admin)
        r = requests.post(f"{BASE}/api/projects/{pub_id}/testcases", headers=h, json={
            "no": 12001, "tc_id": "TC-SRCH-FIND-001", "test_steps": "search target step"
        })
        tc_db_id = r.json().get("id") if r.status_code == 201 else None

        r2 = requests.get(f"{BASE}/api/search?q=TC-SRCH-FIND-001", headers=h)
        assert r2.status_code == 200
        tc_ids = [tc["tc_id"] for tc in r2.json()]
        assert "TC-SRCH-FIND-001" in tc_ids

        # cleanup
        if tc_db_id:
            requests.delete(f"{BASE}/api/projects/{pub_id}/testcases/{tc_db_id}", headers=h)

    def test_search_empty_query(self):
        """빈 쿼리로 검색 시 422 (min_length=1 위반)"""
        h = auth(store.admin)
        r = requests.get(f"{BASE}/api/search?q=", headers=h)
        assert r.status_code == 422

    def test_search_no_match(self):
        """없는 키워드 검색 시 빈 배열"""
        h = auth(store.admin)
        r = requests.get(f"{BASE}/api/search?q=__absolutely_no_match_xyz_99999__", headers=h)
        assert r.status_code == 200
        assert r.json() == []

    def test_search_requires_auth(self):
        """인증 없이 검색 시 401"""
        r = requests.get(f"{BASE}/api/search?q=test")
        assert r.status_code == 401

    def test_search_case_insensitive(self, project_pair):
        """대소문자 무관 검색 (TC ID가 대문자인데 소문자로 검색)"""
        pub_id, _ = project_pair
        h = auth(store.admin)
        r = requests.post(f"{BASE}/api/projects/{pub_id}/testcases", headers=h, json={
            "no": 12002, "tc_id": "TC-SRCH-CASE-UPPER", "test_steps": "case test"
        })
        tc_db_id = r.json().get("id") if r.status_code == 201 else None

        r2 = requests.get(f"{BASE}/api/search?q=tc-srch-case-upper", headers=h)
        assert r2.status_code == 200
        tc_ids = [tc["tc_id"] for tc in r2.json()]
        assert "TC-SRCH-CASE-UPPER" in tc_ids

        # cleanup
        if tc_db_id:
            requests.delete(f"{BASE}/api/projects/{pub_id}/testcases/{tc_db_id}", headers=h)

    def test_search_special_characters(self):
        """특수문자 포함 검색 시 에러 없이 처리"""
        h = auth(store.admin)
        r = requests.get(f"{BASE}/api/search?q=%25DROP+TABLE", headers=h)
        assert r.status_code == 200
        assert isinstance(r.json(), list)


# ── 13. 변경 이력 API ─────────────────────────────────────────


class TestHistory:
    """GET /api/history 엔드포인트 테스트 (5개)."""

    def test_project_history(self, project_pair):
        """GET /api/history/project/{id} 정상 조회"""
        pub_id, _ = project_pair
        h = auth(store.admin)
        r = requests.get(f"{BASE}/api/history/project/{pub_id}", headers=h)
        assert r.status_code == 200
        assert isinstance(r.json(), list)

    def test_project_history_requires_auth(self, project_pair):
        """인증 없이 프로젝트 이력 조회 시 401"""
        pub_id, _ = project_pair
        r = requests.get(f"{BASE}/api/history/project/{pub_id}")
        assert r.status_code == 401

    def test_testcase_history(self, project_pair):
        """GET /api/history/testcase/{tc_id} 정상 조회"""
        pub_id, _ = project_pair
        h = auth(store.admin)
        # TC 생성
        r = requests.post(f"{BASE}/api/projects/{pub_id}/testcases", headers=h, json={
            "no": 13001, "tc_id": "TC-HIST-001", "test_steps": "history test"
        })
        tc_db_id = r.json().get("id") if r.status_code == 201 else None

        if tc_db_id:
            r2 = requests.get(f"{BASE}/api/history/testcase/{tc_db_id}", headers=h)
            assert r2.status_code == 200
            assert isinstance(r2.json(), list)

        # cleanup
        if tc_db_id:
            requests.delete(f"{BASE}/api/projects/{pub_id}/testcases/{tc_db_id}", headers=h)

    def test_testcase_history_empty(self, project_pair):
        """변경 이력 없는 TC 조회 시 빈 배열"""
        pub_id, _ = project_pair
        h = auth(store.admin)
        # TC 생성 직후 (수정 없음) → 이력 비어 있어야 함
        r = requests.post(f"{BASE}/api/projects/{pub_id}/testcases", headers=h, json={
            "no": 13002, "tc_id": "TC-HIST-EMPTY-001", "test_steps": "no changes"
        })
        tc_db_id = r.json().get("id") if r.status_code == 201 else None

        if tc_db_id:
            r2 = requests.get(f"{BASE}/api/history/testcase/{tc_db_id}", headers=h)
            assert r2.status_code == 200
            assert r2.json() == []

        # cleanup
        if tc_db_id:
            requests.delete(f"{BASE}/api/projects/{pub_id}/testcases/{tc_db_id}", headers=h)

    def test_history_after_update(self, project_pair):
        """TC 수정 후 이력에 변경 내용이 기록되는지"""
        pub_id, _ = project_pair
        h = auth(store.admin)
        # TC 생성
        r = requests.post(f"{BASE}/api/projects/{pub_id}/testcases", headers=h, json={
            "no": 13003, "tc_id": "TC-HIST-UPD-001", "test_steps": "original step"
        })
        tc_db_id = r.json().get("id") if r.status_code == 201 else None
        assert tc_db_id, "TC creation failed"

        # TC 수정
        requests.put(f"{BASE}/api/projects/{pub_id}/testcases/{tc_db_id}", headers=h, json={
            "test_steps": "updated step"
        })

        # 이력 확인
        r2 = requests.get(f"{BASE}/api/history/testcase/{tc_db_id}", headers=h)
        assert r2.status_code == 200
        history = r2.json()
        assert len(history) >= 1
        fields_changed = [h["field_name"] for h in history]
        assert "test_steps" in fields_changed

        # cleanup
        requests.delete(f"{BASE}/api/projects/{pub_id}/testcases/{tc_db_id}", headers=h)


# ── 14. 엣지 케이스 ───────────────────────────────────────────


class TestEdgeCases:
    """기존 엔드포인트의 엣지 케이스 테스트 (5개)."""

    def test_create_tc_missing_required_fields(self, project_pair):
        """TC 생성 시 필수 필드(no, tc_id) 없으면 422"""
        pub_id, _ = project_pair
        h = auth(store.admin)
        # no, tc_id 모두 빠진 요청
        r = requests.post(f"{BASE}/api/projects/{pub_id}/testcases", headers=h, json={
            "test_steps": "only steps"
        })
        assert r.status_code == 422

    def test_update_nonexistent_tc(self, project_pair):
        """존재하지 않는 TC 수정 시 404"""
        pub_id, _ = project_pair
        h = auth(store.admin)
        r = requests.put(f"{BASE}/api/projects/{pub_id}/testcases/999999", headers=h, json={
            "test_steps": "ghost"
        })
        assert r.status_code == 404

    def test_delete_nonexistent_tc(self, project_pair):
        """존재하지 않는 TC 삭제 시 404"""
        pub_id, _ = project_pair
        h = auth(store.admin)
        r = requests.delete(f"{BASE}/api/projects/{pub_id}/testcases/999999", headers=h)
        assert r.status_code == 404

    def test_create_project_empty_name(self):
        """빈 이름으로 프로젝트 생성 시 422"""
        h = auth(store.admin)
        r = requests.post(f"{BASE}/api/projects", headers=h, json={
            "name": "", "is_private": False
        })
        assert r.status_code == 422

    def test_get_nonexistent_project(self):
        """존재하지 않는 프로젝트 조회 시 404"""
        h = auth(store.admin)
        r = requests.get(f"{BASE}/api/projects/999999", headers=h)
        assert r.status_code == 404

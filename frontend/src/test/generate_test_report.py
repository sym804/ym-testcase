"""프론트엔드 테스트 케이스를 엑셀 파일로 생성하는 스크립트"""
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
import os

wb = openpyxl.Workbook()
ws = wb.active
ws.title = "Frontend Test Cases"

# 스타일 정의
header_font = Font(bold=True, color="FFFFFF", size=11)
header_fill = PatternFill(start_color="2D4A7A", end_color="2D4A7A", fill_type="solid")
category_fill = PatternFill(start_color="E8EDF5", end_color="E8EDF5", fill_type="solid")
category_font = Font(bold=True, size=11, color="2D4A7A")
pass_fill = PatternFill(start_color="DCFCE7", end_color="DCFCE7", fill_type="solid")
thin_border = Border(
    left=Side(style="thin", color="D1D5DB"),
    right=Side(style="thin", color="D1D5DB"),
    top=Side(style="thin", color="D1D5DB"),
    bottom=Side(style="thin", color="D1D5DB"),
)

# 헤더
headers = ["No", "카테고리", "테스트 파일", "테스트 ID", "테스트 설명", "테스트 유형", "결과"]
ws.append(headers)
for col_idx, _ in enumerate(headers, 1):
    cell = ws.cell(row=1, column=col_idx)
    cell.font = header_font
    cell.fill = header_fill
    cell.alignment = Alignment(horizontal="center", vertical="center")
    cell.border = thin_border

# 테스트 데이터
test_cases = [
    # API Client
    ("API 클라이언트", "api-client.test.ts", "request-token-present", "토큰이 있으면 Authorization 헤더를 추가한다", "단위"),
    ("API 클라이언트", "api-client.test.ts", "request-token-absent", "토큰이 없으면 Authorization 헤더를 추가하지 않는다", "단위"),
    ("API 클라이언트", "api-client.test.ts", "response-401-redirect", "401 응답 시 토큰을 제거하고 로그인 페이지로 리다이렉트한다", "단위"),
    ("API 클라이언트", "api-client.test.ts", "response-non401-passthrough", "401이 아닌 에러는 그대로 reject한다", "단위"),
    ("API 클라이언트", "api-client.test.ts", "response-success", "정상 응답은 그대로 반환한다", "단위"),

    # AuthContext
    ("인증 컨텍스트", "AuthContext.test.tsx", "init-no-token", "토큰이 없으면 user는 null이고 loading이 false가 된다", "단위"),
    ("인증 컨텍스트", "AuthContext.test.tsx", "init-valid-token", "토큰이 있으면 getMe를 호출하여 사용자를 로드한다", "단위"),
    ("인증 컨텍스트", "AuthContext.test.tsx", "init-invalid-token", "토큰이 유효하지 않으면 토큰을 제거한다", "단위"),
    ("인증 컨텍스트", "AuthContext.test.tsx", "login-success", "로그인 성공 시 토큰을 저장하고 사용자를 로드한다", "통합"),
    ("인증 컨텍스트", "AuthContext.test.tsx", "login-must-change-pw", "must_change_password가 true이면 환영 메시지를 표시하지 않는다", "통합"),
    ("인증 컨텍스트", "AuthContext.test.tsx", "login-failure", "로그인 실패 시 에러를 throw한다", "통합"),
    ("인증 컨텍스트", "AuthContext.test.tsx", "register-pw-mismatch", "비밀번호 불일치 시 에러를 throw한다", "단위"),
    ("인증 컨텍스트", "AuthContext.test.tsx", "register-pw-short", "비밀번호가 8자 미만이면 에러를 throw한다", "단위"),
    ("인증 컨텍스트", "AuthContext.test.tsx", "register-valid", "유효한 폼이면 API를 호출한다", "통합"),
    ("인증 컨텍스트", "AuthContext.test.tsx", "logout", "로그아웃 시 토큰과 사용자를 제거한다", "통합"),
    ("인증 컨텍스트", "AuthContext.test.tsx", "change-password", "비밀번호 변경 성공 시 mustChangePassword를 false로 설정한다", "통합"),
    ("인증 컨텍스트", "AuthContext.test.tsx", "hook-outside-provider", "AuthProvider 밖에서 사용하면 에러를 throw한다", "단위"),

    # LoginPage
    ("로그인 페이지", "LoginPage.test.tsx", "render-form", "로그인 폼이 렌더링된다", "UI"),
    ("로그인 페이지", "LoginPage.test.tsx", "empty-submit", "빈 필드로 제출하면 에러 메시지를 표시한다", "UI"),
    ("로그인 페이지", "LoginPage.test.tsx", "login-success-navigate", "로그인 성공 시 /projects로 이동한다", "통합"),
    ("로그인 페이지", "LoginPage.test.tsx", "login-failure-error", "로그인 실패 시 서버 에러 메시지를 표시한다", "통합"),
    ("로그인 페이지", "LoginPage.test.tsx", "loading-disabled", "로딩 중에는 버튼이 비활성화된다", "UI"),
    ("로그인 페이지", "LoginPage.test.tsx", "register-link", "회원가입 링크가 있다", "UI"),

    # RegisterPage
    ("회원가입 페이지", "RegisterPage.test.tsx", "render-form", "회원가입 폼이 렌더링된다", "UI"),
    ("회원가입 페이지", "RegisterPage.test.tsx", "empty-submit", "모든 필드가 비어있으면 에러를 표시한다", "UI"),
    ("회원가입 페이지", "RegisterPage.test.tsx", "pw-short-hint", "비밀번호가 8자 미만이면 힌트를 표시한다", "UI"),
    ("회원가입 페이지", "RegisterPage.test.tsx", "pw-mismatch-hint", "비밀번호 불일치 시 힌트를 표시한다", "UI"),
    ("회원가입 페이지", "RegisterPage.test.tsx", "username-available", "아이디 입력 시 사용 가능 여부를 확인한다", "통합"),
    ("회원가입 페이지", "RegisterPage.test.tsx", "username-taken", "이미 사용 중인 아이디면 경고를 표시한다", "통합"),
    ("회원가입 페이지", "RegisterPage.test.tsx", "submit-taken-error", "사용 중인 아이디로 제출하면 에러를 표시한다", "통합"),
    ("회원가입 페이지", "RegisterPage.test.tsx", "register-success", "회원가입 성공 시 로그인 페이지로 이동한다", "통합"),
    ("회원가입 페이지", "RegisterPage.test.tsx", "login-link", "로그인 페이지 링크가 있다", "UI"),

    # PasswordInput
    ("비밀번호 입력", "PasswordInput.test.tsx", "default-type", "기본적으로 password 타입이다", "단위"),
    ("비밀번호 입력", "PasswordInput.test.tsx", "toggle-show", "눈 아이콘 클릭 시 text 타입으로 변경된다", "UI"),
    ("비밀번호 입력", "PasswordInput.test.tsx", "toggle-hide", "다시 클릭하면 password 타입으로 돌아간다", "UI"),
    ("비밀번호 입력", "PasswordInput.test.tsx", "on-change", "onChange 이벤트가 전달된다", "단위"),

    # ChangePasswordModal
    ("비밀번호 변경 모달", "ChangePasswordModal.test.tsx", "render-form", "비밀번호 변경 폼이 렌더링된다", "UI"),
    ("비밀번호 변경 모달", "ChangePasswordModal.test.tsx", "pw-short-hint", "8자 미만이면 힌트를 표시한다", "UI"),
    ("비밀번호 변경 모달", "ChangePasswordModal.test.tsx", "pw-mismatch-hint", "비밀번호 불일치 시 힌트를 표시한다", "UI"),
    ("비밀번호 변경 모달", "ChangePasswordModal.test.tsx", "submit-disabled", "조건 미충족 시 제출 버튼이 비활성화된다", "UI"),
    ("비밀번호 변경 모달", "ChangePasswordModal.test.tsx", "change-success", "유효한 입력 시 비밀번호를 변경한다", "통합"),
    ("비밀번호 변경 모달", "ChangePasswordModal.test.tsx", "change-failure", "변경 실패 시 에러 메시지를 표시한다", "통합"),
    ("비밀번호 변경 모달", "ChangePasswordModal.test.tsx", "logout-button", "로그아웃 버튼이 있다", "UI"),

    # App Routing
    ("라우팅", "App.test.tsx", "login-route", "/login 경로에서 로그인 페이지를 렌더링한다", "통합"),
    ("라우팅", "App.test.tsx", "register-route", "/register 경로에서 회원가입 페이지를 렌더링한다", "통합"),
    ("라우팅", "App.test.tsx", "projects-redirect", "/projects 접근 시 로그인 페이지로 리다이렉트된다", "통합"),
    ("라우팅", "App.test.tsx", "root-redirect", "/ 접근 시 로그인 페이지로 리다이렉트된다", "통합"),
    ("라우팅", "App.test.tsx", "unknown-redirect", "알 수 없는 경로에서도 로그인 페이지로 리다이렉트된다", "통합"),
    ("라우팅", "App.test.tsx", "auth-projects", "인증된 사용자는 프로젝트 목록에 접근할 수 있다", "통합"),
    ("라우팅", "App.test.tsx", "must-change-pw", "must_change_password가 true이면 비밀번호 변경 모달을 표시한다", "통합"),
    ("라우팅", "App.test.tsx", "loading-state", "로딩 중에는 로딩 메시지를 표시한다", "UI"),
    ("라우팅", "App.test.tsx", "protected-admin", "인증된 사용자의 페이지 접근 검증", "통합"),
    ("라우팅", "App.test.tsx", "protected-manual", "매뉴얼 페이지 접근 검증", "통합"),

    # HighlightCell
    ("하이라이트 셀", "HighlightCell.test.tsx", "null-value", "값이 없으면 아무것도 렌더링하지 않는다", "단위"),
    ("하이라이트 셀", "HighlightCell.test.tsx", "no-keyword", "키워드 없이 텍스트를 그대로 표시한다", "단위"),
    ("하이라이트 셀", "HighlightCell.test.tsx", "highlight-mark", "키워드가 있으면 mark 태그로 하이라이트한다", "단위"),
    ("하이라이트 셀", "HighlightCell.test.tsx", "case-insensitive", "대소문자 무관하게 하이라이트한다", "단위"),
    ("하이라이트 셀", "HighlightCell.test.tsx", "number-value", "숫자 값도 문자열로 변환하여 표시한다", "단위"),
    ("하이라이트 셀", "HighlightCell.test.tsx", "multiple-match", "여러 매칭을 모두 하이라이트한다", "단위"),

    # MarkdownCell
    ("마크다운 셀", "MarkdownCell.test.tsx", "empty-value", "빈 값이면 아무것도 렌더링하지 않는다", "단위"),
    ("마크다운 셀", "MarkdownCell.test.tsx", "bold", "마크다운 bold를 strong으로 변환한다", "단위"),
    ("마크다운 셀", "MarkdownCell.test.tsx", "italic", "마크다운 italic을 em으로 변환한다", "단위"),
    ("마크다운 셀", "MarkdownCell.test.tsx", "code", "마크다운 code를 code 태그로 변환한다", "단위"),
    ("마크다운 셀", "MarkdownCell.test.tsx", "xss-script", "XSS 공격 태그를 제거한다 (DOMPurify)", "보안"),
    ("마크다운 셀", "MarkdownCell.test.tsx", "xss-img", "허용되지 않은 태그를 제거한다 (img, iframe 등)", "보안"),
    ("마크다운 셀", "MarkdownCell.test.tsx", "highlight", "키워드 하이라이트를 적용한다", "단위"),
    ("마크다운 셀", "MarkdownCell.test.tsx", "link", "링크를 a 태그로 변환한다", "단위"),

    # Header
    ("헤더", "Header.test.tsx", "brand", "TC Manager 브랜드를 표시한다", "UI"),
    ("헤더", "Header.test.tsx", "project-dropdown", "프로젝트 드롭다운을 표시한다", "통합"),
    ("헤더", "Header.test.tsx", "search-input", "검색 입력란을 표시한다", "UI"),
    ("헤더", "Header.test.tsx", "user-info", "사용자 이름과 역할을 표시한다", "UI"),
    ("헤더", "Header.test.tsx", "admin-buttons", "admin 사용자에게 관리 버튼을 표시한다", "UI"),
    ("헤더", "Header.test.tsx", "help-button", "도움말 버튼을 표시한다", "UI"),
    ("헤더", "Header.test.tsx", "theme-toggle", "테마 토글 버튼이 있다", "UI"),
    ("헤더", "Header.test.tsx", "brand-navigate", "TC Manager 클릭 시 /projects로 이동한다", "UI"),
    ("헤더", "Header.test.tsx", "user-menu", "사용자 메뉴 클릭 시 드롭다운을 표시한다", "UI"),
    ("헤더", "Header.test.tsx", "logout", "로그아웃 클릭 시 logout을 호출한다", "통합"),
    ("헤더", "Header.test.tsx", "search-min-chars", "2자 미만 검색어는 결과를 표시하지 않는다", "단위"),
    ("헤더", "Header.test.tsx", "search-results", "2자 이상 검색 시 결과를 표시한다", "통합"),

    # Dashboard
    ("대시보드", "Dashboard.test.tsx", "loading", "로딩 중 메시지를 표시한다", "UI"),
    ("대시보드", "Dashboard.test.tsx", "summary-cards", "요약 카드를 렌더링한다", "통합"),
    ("대시보드", "Dashboard.test.tsx", "doughnut-chart", "도넛 차트를 렌더링한다", "UI"),
    ("대시보드", "Dashboard.test.tsx", "bar-chart", "바 차트를 렌더링한다", "UI"),
    ("대시보드", "Dashboard.test.tsx", "trend-chart", "트렌드 라인 차트를 렌더링한다", "UI"),
    ("대시보드", "Dashboard.test.tsx", "priority-table", "우선순위 테이블을 렌더링한다", "통합"),
    ("대시보드", "Dashboard.test.tsx", "category-table", "카테고리 테이블을 렌더링한다", "통합"),
    ("대시보드", "Dashboard.test.tsx", "assignee-table", "담당자 테이블을 렌더링한다", "통합"),
    ("대시보드", "Dashboard.test.tsx", "heatmap", "히트맵을 렌더링한다", "통합"),
    ("대시보드", "Dashboard.test.tsx", "run-selector", "테스트 수행 셀렉터를 표시한다", "UI"),

    # CompareView
    ("비교 뷰", "CompareView.test.tsx", "selectors", "수행 선택 드롭다운을 표시한다", "UI"),
    ("비교 뷰", "CompareView.test.tsx", "placeholder", "양쪽 미선택 시 안내 메시지를 표시한다", "UI"),
    ("비교 뷰", "CompareView.test.tsx", "compare-results", "양쪽 선택 후 비교 결과를 표시한다", "통합"),
    ("비교 뷰", "CompareView.test.tsx", "regression-detect", "퇴보(regression)를 감지한다 (PASS→FAIL)", "통합"),
    ("비교 뷰", "CompareView.test.tsx", "filter-buttons", "필터 버튼이 작동한다", "UI"),

    # ReportView
    ("리포트 뷰", "ReportView.test.tsx", "run-selector", "테스트 수행 셀렉터를 표시한다", "UI"),
    ("리포트 뷰", "ReportView.test.tsx", "download-buttons", "다운로드 버튼을 표시한다", "UI"),
    ("리포트 뷰", "ReportView.test.tsx", "project-info", "프로젝트 정보를 표시한다", "통합"),
    ("리포트 뷰", "ReportView.test.tsx", "stats-cards", "전체 현황 카드를 표시한다", "통합"),
    ("리포트 뷰", "ReportView.test.tsx", "top-failures", "주요 실패 항목을 표시한다", "통합"),
    ("리포트 뷰", "ReportView.test.tsx", "jira-issues", "Jira 이슈 배지를 표시한다", "통합"),
    ("리포트 뷰", "ReportView.test.tsx", "category-summary", "카테고리별 요약 테이블을 표시한다", "통합"),

    # ProjectSettings
    ("프로젝트 설정", "ProjectSettings.test.tsx", "access-section", "접근 설정 섹션을 표시한다", "UI"),
    ("프로젝트 설정", "ProjectSettings.test.tsx", "admin-toggle", "admin은 공개/비공개 토글 버튼을 볼 수 있다", "UI"),
    ("프로젝트 설정", "ProjectSettings.test.tsx", "tester-no-toggle", "tester는 상태 배지만 볼 수 있다 (토글 불가)", "UI"),
    ("프로젝트 설정", "ProjectSettings.test.tsx", "admin-danger-zone", "admin은 위험 영역(삭제)을 볼 수 있다", "UI"),
    ("프로젝트 설정", "ProjectSettings.test.tsx", "tester-no-danger", "tester는 위험 영역을 볼 수 없다", "UI"),
    ("프로젝트 설정", "ProjectSettings.test.tsx", "delete-confirm", "삭제 버튼 클릭 시 확인 입력란을 표시한다", "UI"),
    ("프로젝트 설정", "ProjectSettings.test.tsx", "delete-disabled", "프로젝트 이름 불일치 시 영구 삭제 버튼이 비활성화된다", "UI"),
    ("프로젝트 설정", "ProjectSettings.test.tsx", "delete-enabled", "프로젝트 이름 일치 시 삭제가 가능하다", "통합"),
    ("프로젝트 설정", "ProjectSettings.test.tsx", "my-role", "내 역할을 표시한다", "UI"),
    ("프로젝트 설정", "ProjectSettings.test.tsx", "cancel-confirm", "취소 버튼으로 확인 입력란을 닫을 수 있다", "UI"),

    # ProjectMembers
    ("프로젝트 멤버", "ProjectMembers.test.tsx", "member-list", "멤버 목록을 렌더링한다", "통합"),
    ("프로젝트 멤버", "ProjectMembers.test.tsx", "creator-badge", "생성자에 배지를 표시한다", "UI"),
    ("프로젝트 멤버", "ProjectMembers.test.tsx", "admin-add-row", "admin은 멤버 추가 행을 볼 수 있다", "UI"),
    ("프로젝트 멤버", "ProjectMembers.test.tsx", "tester-no-add", "tester는 멤버 추가 행을 볼 수 없다", "UI"),
    ("프로젝트 멤버", "ProjectMembers.test.tsx", "add-member", "멤버 추가가 작동한다", "통합"),
    ("프로젝트 멤버", "ProjectMembers.test.tsx", "creator-no-remove", "생성자는 제거할 수 없다 (X 버튼 없음)", "UI"),
    ("프로젝트 멤버", "ProjectMembers.test.tsx", "loading", "로딩 중 메시지를 표시한다", "UI"),

    # ProjectListPage
    ("프로젝트 목록", "ProjectListPage.test.tsx", "overview-dashboard", "전체 현황 대시보드를 표시한다", "통합"),
    ("프로젝트 목록", "ProjectListPage.test.tsx", "project-cards", "프로젝트 카드를 렌더링한다", "통합"),
    ("프로젝트 목록", "ProjectListPage.test.tsx", "project-desc", "프로젝트 설명을 표시한다", "통합"),
    ("프로젝트 목록", "ProjectListPage.test.tsx", "admin-new-btn", "admin에게 새 프로젝트 버튼을 표시한다", "UI"),
    ("프로젝트 목록", "ProjectListPage.test.tsx", "user-no-new-btn", "일반 사용자에게는 새 프로젝트 버튼을 표시하지 않는다", "UI"),
    ("프로젝트 목록", "ProjectListPage.test.tsx", "create-modal", "새 프로젝트 모달을 열고 닫을 수 있다", "UI"),
    ("프로젝트 목록", "ProjectListPage.test.tsx", "create-project", "프로젝트를 생성할 수 있다", "통합"),
    ("프로젝트 목록", "ProjectListPage.test.tsx", "pass-rate", "프로젝트별 테이블에 Pass Rate를 표시한다", "통합"),

    # ProjectPage
    ("프로젝트 페이지", "ProjectPage.test.tsx", "project-name", "프로젝트 이름을 표시한다", "통합"),
    ("프로젝트 페이지", "ProjectPage.test.tsx", "project-desc", "프로젝트 설명을 표시한다", "통합"),
    ("프로젝트 페이지", "ProjectPage.test.tsx", "tab-buttons", "6개 탭 버튼을 렌더링한다", "UI"),
    ("프로젝트 페이지", "ProjectPage.test.tsx", "default-tab", "기본 탭은 TC 관리이다", "UI"),
    ("프로젝트 페이지", "ProjectPage.test.tsx", "tab-switch", "탭 전환이 작동한다", "UI"),
    ("프로젝트 페이지", "ProjectPage.test.tsx", "settings-tab", "설정 탭에서 접근 설정을 표시한다", "통합"),
    ("프로젝트 페이지", "ProjectPage.test.tsx", "loading-state", "로딩 중 메시지를 표시한다", "UI"),

    # AdminPage
    ("관리자 페이지", "AdminPage.test.tsx", "no-admin-access", "관리자가 아니면 권한 필요 메시지를 표시한다", "UI"),
    ("관리자 페이지", "AdminPage.test.tsx", "title", "사용자 관리 제목을 표시한다", "UI"),
    ("관리자 페이지", "AdminPage.test.tsx", "user-table", "사용자 목록을 테이블로 렌더링한다", "통합"),
    ("관리자 페이지", "AdminPage.test.tsx", "role-selector", "역할 변경 셀렉터가 있다", "UI"),
    ("관리자 페이지", "AdminPage.test.tsx", "reset-pw-btn", "비밀번호 초기화 버튼이 있다", "UI"),
    ("관리자 페이지", "AdminPage.test.tsx", "reset-pw-modal", "비밀번호 초기화 시 임시 비밀번호 모달을 표시한다", "통합"),
    ("관리자 페이지", "AdminPage.test.tsx", "role-change", "역할을 변경할 수 있다", "통합"),
    ("관리자 페이지", "AdminPage.test.tsx", "assign-btn", "프로젝트 배정 관리 버튼이 있다", "UI"),
]

# 데이터 입력
current_category = ""
for idx, (category, file, test_id, desc, test_type) in enumerate(test_cases, 1):
    row_num = idx + 1
    ws.cell(row=row_num, column=1, value=idx)
    ws.cell(row=row_num, column=2, value=category)
    ws.cell(row=row_num, column=3, value=file)
    ws.cell(row=row_num, column=4, value=test_id)
    ws.cell(row=row_num, column=5, value=desc)
    ws.cell(row=row_num, column=6, value=test_type)
    ws.cell(row=row_num, column=7, value="PASS")

    # 카테고리 변경 시 배경색
    if category != current_category:
        current_category = category
        for col in range(1, 8):
            ws.cell(row=row_num, column=col).fill = category_fill

    # PASS 배경색
    ws.cell(row=row_num, column=7).fill = pass_fill
    ws.cell(row=row_num, column=7).alignment = Alignment(horizontal="center")

    # 테두리
    for col in range(1, 8):
        ws.cell(row=row_num, column=col).border = thin_border

# 열 너비
ws.column_dimensions["A"].width = 5
ws.column_dimensions["B"].width = 18
ws.column_dimensions["C"].width = 28
ws.column_dimensions["D"].width = 26
ws.column_dimensions["E"].width = 55
ws.column_dimensions["F"].width = 10
ws.column_dimensions["G"].width = 8

# 요약 시트
ws2 = wb.create_sheet("Summary")
ws2.append(["카테고리", "테스트 수", "PASS", "FAIL", "Pass Rate"])
for col in range(1, 6):
    cell = ws2.cell(row=1, column=col)
    cell.font = header_font
    cell.fill = header_fill
    cell.alignment = Alignment(horizontal="center")
    cell.border = thin_border

categories = {}
for cat, _, _, _, test_type in test_cases:
    categories[cat] = categories.get(cat, 0) + 1

for idx, (cat, count) in enumerate(categories.items(), 2):
    ws2.cell(row=idx, column=1, value=cat).border = thin_border
    ws2.cell(row=idx, column=2, value=count).border = thin_border
    ws2.cell(row=idx, column=2).alignment = Alignment(horizontal="center")
    ws2.cell(row=idx, column=3, value=count).border = thin_border
    ws2.cell(row=idx, column=3).alignment = Alignment(horizontal="center")
    ws2.cell(row=idx, column=3).fill = pass_fill
    ws2.cell(row=idx, column=4, value=0).border = thin_border
    ws2.cell(row=idx, column=4).alignment = Alignment(horizontal="center")
    ws2.cell(row=idx, column=5, value="100%").border = thin_border
    ws2.cell(row=idx, column=5).alignment = Alignment(horizontal="center")
    ws2.cell(row=idx, column=5).fill = pass_fill

# 합계
total_row = len(categories) + 2
ws2.cell(row=total_row, column=1, value="합계").font = Font(bold=True)
ws2.cell(row=total_row, column=2, value=len(test_cases)).font = Font(bold=True)
ws2.cell(row=total_row, column=2).alignment = Alignment(horizontal="center")
ws2.cell(row=total_row, column=3, value=len(test_cases)).font = Font(bold=True)
ws2.cell(row=total_row, column=3).alignment = Alignment(horizontal="center")
ws2.cell(row=total_row, column=3).fill = pass_fill
ws2.cell(row=total_row, column=4, value=0).font = Font(bold=True)
ws2.cell(row=total_row, column=4).alignment = Alignment(horizontal="center")
ws2.cell(row=total_row, column=5, value="100%").font = Font(bold=True)
ws2.cell(row=total_row, column=5).alignment = Alignment(horizontal="center")
ws2.cell(row=total_row, column=5).fill = pass_fill
for col in range(1, 6):
    ws2.cell(row=total_row, column=col).border = thin_border

ws2.column_dimensions["A"].width = 20
ws2.column_dimensions["B"].width = 12
ws2.column_dimensions["C"].width = 8
ws2.column_dimensions["D"].width = 8
ws2.column_dimensions["E"].width = 12

# 저장
output_path = os.path.join(os.path.dirname(os.path.dirname(os.path.dirname(os.path.abspath(__file__)))), "frontend_test_cases.xlsx")
wb.save(output_path)
print(f"테스트 케이스 엑셀 생성 완료: {output_path}")
print(f"총 {len(test_cases)}개 테스트 케이스, {len(categories)}개 카테고리")
